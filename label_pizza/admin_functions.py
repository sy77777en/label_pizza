import streamlit as st
import pandas as pd
from typing import Dict, Optional, List, Any, Tuple
from sqlalchemy.orm import Session

from label_pizza.services import (
    AuthService, GroundTruthService, 
    QuestionService, QuestionGroupService, SchemaService,
    ProjectService, VideoService, ProjectGroupService
)
import label_pizza.export as export_module
from label_pizza.ui_components import (
    custom_info, get_card_style, COLORS
)
from label_pizza.database_utils import (
    get_db_session, check_project_has_full_ground_truth, 
    handle_database_errors, get_project_groups_with_projects,
    get_annotator_accuracy_cached, get_reviewer_accuracy_cached,
    get_cached_project_metadata
)
from label_pizza.accuracy_analytics import (
    get_accuracy_color, calculate_overall_accuracy
)

###############################################################################
# ADMIN PORTAL - OPTIMIZED WITH FRAGMENTS
###############################################################################

@handle_database_errors
def admin_portal():
    st.title("⚙️ Admin Portal")
    
    tabs = st.tabs([
        "📹 Videos", "📁 Projects", "📋 Schemas", 
        "❓ Questions", "👥 Users", "🔗 Assignments", "📊 Project Groups", "📤 Export"
    ])
    
    with tabs[0]:
        admin_videos()
    with tabs[1]:
        admin_projects()
    with tabs[2]:
        admin_schemas()
    with tabs[3]:
        admin_questions()
    with tabs[4]:
        admin_users()
    with tabs[5]:
        admin_assignments()
    with tabs[6]:
        admin_project_groups()
    with tabs[7]:
        admin_export()


@st.fragment
def admin_videos():
    st.subheader("📹 Video Management")
    
    # Load only counts for metrics - much faster
    try:
        with get_db_session() as session:
            video_counts = VideoService.get_video_counts(session=session)
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📹 Total Videos", video_counts["total"])
        with col2:
            st.metric("✅ Active Videos", video_counts["active"])
        with col3:
            st.metric("🗄️ Archived Videos", video_counts["archived"])
        with col4:
            st.metric("📭 Unassigned Videos", video_counts["unassigned"])
            
    except Exception as e:
        st.error(f"Error loading video summary: {str(e)}")
        video_counts = {"total": 0}
    
    # Option to view full video table - now as a dialog
    if video_counts["total"] > 0:
        if st.button("📋 View All Videos (Database Table)", use_container_width=True):
            clear_other_dialogs('videos')
            st.rerun()
        
        if st.session_state.get('show_videos_dialog', False):
            st.session_state['show_videos_dialog'] = False
            show_videos_table_dialog()
    
    video_tabs = st.tabs(["➕ Add Video", "✏️ Edit Video"])
    
    with video_tabs[0]:
        video_uid = st.text_input("Video UID", key="admin_video_uid")
        url = st.text_input("Video URL", key="admin_video_url")
        metadata_json = st.text_area("Metadata (JSON, optional)", "{}", key="admin_video_metadata")
        
        if st.button("Add Video", key="admin_add_video_btn"):
            if url:
                try:
                    import json
                    metadata = json.loads(metadata_json) if metadata_json.strip() else {}
                    with get_db_session() as session:
                        VideoService.add_video(video_uid=video_uid, url=url, session=session, metadata=metadata)
                    custom_info("Video added!")
                    # st.rerun(scope="fragment")
                except Exception as e:
                    st.error(f"Error: {str(e)}")
    
    with video_tabs[1]:
        show_video_edit_interface()


def show_video_edit_interface():
    """Show video editing interface with search-first approach."""
    st.markdown("### ✏️ Edit Video")
    
    # Search-first approach
    search_term = st.text_input(
        "🔍 Search for video to edit", 
        placeholder="Enter video UID to search...",
        key="admin_edit_video_search",
        help="Enter at least 3 characters to search"
    )
    
    if len(search_term) >= 3:
        try:
            # Search for videos matching the term
            with get_db_session() as session:
                matching_videos = VideoService.search_videos_for_selection(
                    search_term=search_term, limit=20, session=session
                )
            
            if matching_videos:
                if len(matching_videos) == 20:
                    st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term.")
                else:
                    custom_info(f"✅ Found {len(matching_videos)} matching videos")
                
                # Show matching videos for selection
                video_options = {f"{v['uid']} - {v['url']}": v['uid'] for v in matching_videos}
                
                selected_video_display = st.selectbox(
                    f"Select from {len(matching_videos)} matching videos",
                    [""] + list(video_options.keys()),
                    key="admin_edit_video_select"
                )
                
                if selected_video_display:
                    selected_video_uid = video_options[selected_video_display]
                    show_video_edit_form(selected_video_uid)
            else:
                custom_info(f"No videos found matching '{search_term}'")
        except Exception as e:
            st.error(f"Error searching videos: {str(e)}")
    else:
        custom_info("Enter at least 3 characters in the search box to find videos to edit")

def show_video_edit_form(video_uid: str):
    """Show the actual video editing form."""
    try:
        with get_db_session() as session:
            current_video = VideoService.get_video_by_uid(video_uid=video_uid, session=session)
        
        if current_video:
            st.markdown(f"**Editing Video:** {video_uid}")
            
            col1, col2 = st.columns([2, 1])
            
            with col1:
                new_url = st.text_input(
                    "Video URL",
                    value=current_video.url,
                    key="admin_edit_video_url",
                    help="Update the video URL"
                )
            
            with col2:
                st.markdown(f"**Current Video UID:** {video_uid}")
                st.caption("Video UID cannot be changed")
            
            # Metadata editing - JSON Editor Only
            st.markdown("**Video Metadata (JSON Editor):**")
            current_metadata = current_video.video_metadata or {}
            
            import json
            metadata_json = st.text_area(
                "Metadata JSON",
                value=json.dumps(current_metadata, indent=2) if current_metadata else "{}",
                height=300,
                key="admin_edit_video_metadata_json",
                help="Edit video metadata as JSON. Use valid JSON format."
            )
            
            try:
                updated_metadata = json.loads(metadata_json)
                custom_info("✅ Valid JSON")
            except json.JSONDecodeError as e:
                st.error(f"❌ Invalid JSON: {str(e)}")
                updated_metadata = current_metadata
            
            # Update button with preview
            update_col, preview_col = st.columns(2)
            
            with update_col:
                if st.button("💾 Update Video", key="admin_update_video_btn", use_container_width=True):
                    try:
                        changes_made = []
                        
                        # Check what changed
                        if new_url != current_video.url:
                            changes_made.append("URL")
                        
                        if updated_metadata != current_video.video_metadata:
                            changes_made.append("Metadata")
                        
                        if changes_made:
                            with get_db_session() as session:
                                VideoService.update_video(
                                    video_uid=video_uid, 
                                    new_url=new_url, 
                                    new_metadata=updated_metadata, 
                                    session=session
                                )
                            custom_info(f"✅ Video '{video_uid}' updated successfully! Changed: {', '.join(changes_made)}")
                            # st.rerun(scope="fragment")
                        else:
                            custom_info("No changes were made")
                            
                    except Exception as e:
                        st.error(f"❌ Error updating video: {str(e)}")
            
            with preview_col:
                with st.expander("👁️ Preview Changes"):
                    st.markdown("**Changes to be applied:**")
                    
                    if new_url != current_video.url:
                        st.markdown(f"**URL:** {current_video.url} → {new_url}")
                    
                    if updated_metadata != current_video.video_metadata:
                        st.markdown("**Metadata:** Updated")
                    
                    if new_url == current_video.url and updated_metadata == current_video.video_metadata:
                        custom_info("No changes to preview")
        else:
            st.error(f"Video with UID '{video_uid}' not found")
    except Exception as e:
        st.error(f"Error loading video details: {str(e)}")


@st.fragment
def admin_projects():
    st.subheader("📁 Project Management")
    
    # Load only counts for metrics - much faster
    try:
        with get_db_session() as session:
            project_counts = ProjectService.get_project_counts(session=session)
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📁 Total Projects", project_counts["total"])
        with col2:
            st.metric("✅ Active Projects", project_counts["active"])
        with col3:
            st.metric("🗄️ Archived Projects", project_counts["archived"])
        with col4:
            training_annotation_ratio = f"{project_counts['training_mode']}/{project_counts['annotation_mode']}"
            st.metric("🎓/📝 Train/Annotate", training_annotation_ratio)
    except Exception as e:
        st.error(f"Error loading project summary: {str(e)}")
        project_counts = {"total": 0}
    
    
    # Option to view full project table - now as a dialog
    if project_counts["total"] > 0:
        if st.button("📋 View All Projects (Database Table)", use_container_width=True):
            clear_other_dialogs('projects')
            st.rerun()
        
        if st.session_state.get('show_projects_dialog', False):
            st.session_state['show_projects_dialog'] = False
            show_projects_table_dialog()
    
    # Management tabs
    project_management_tabs = st.tabs(["➕ Create Project", "✏️ Edit Project"])
    
    with project_management_tabs[0]:
        show_project_create_form()
    
    with project_management_tabs[1]:
        show_project_edit_interface()


def show_project_edit_interface():
    """Show project editing interface with search-first approach."""
    st.markdown("### ✏️ Edit Project")
    
    # Search-first approach
    search_term = st.text_input(
        "🔍 Search for project to edit", 
        placeholder="Enter project name to search...",
        key="admin_edit_project_search",
        help="Enter at least 3 characters to search"
    )
    
    if len(search_term) >= 3:
        try:
            with get_db_session() as session:
                # Search for projects matching the term
                matching_projects = ProjectService.search_projects_for_selection(
                    search_term=search_term, limit=20, session=session
                )
            
            if matching_projects:
                # Check if we hit the limit and suggest refinement
                if len(matching_projects) == 20:
                    st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term to find your project faster.")
                else:
                    custom_info(f"✅ Found {len(matching_projects)} matching projects")
                
                # Show matching projects for selection
                project_options = {f"{p['name']} (ID: {p['id']})": p['id'] for p in matching_projects}
                
                selected_project_display = st.selectbox(
                    f"Select from {len(matching_projects)} matching projects" + (" (showing first 20)" if len(matching_projects) == 20 else ""),
                    [""] + list(project_options.keys()),
                    key="admin_edit_project_select"
                )
                
                if selected_project_display:
                    selected_project_id = project_options[selected_project_display]
                    show_project_edit_form(selected_project_id)
            else:
                custom_info(f"No projects found matching '{search_term}'. Try a different search term.")
        except Exception as e:
            st.error(f"Error searching projects: {str(e)}")
    else:
        custom_info("Enter at least 3 characters in the search box to find projects to edit")


def show_project_create_form():
    """Show project creation form with optimized video selection."""
    st.markdown("### 🆕 Create New Project")
    
    name = st.text_input("Project Name", key="admin_project_name", placeholder="Enter project name...")
    description = st.text_area("Description", key="admin_project_description", placeholder="Describe the purpose of this project...")
    
    # Schema selection (keeping this simple since schemas are fewer)
    try:
        with get_db_session() as session:
            schemas_df = SchemaService.get_all_schemas(session=session)
            if schemas_df.empty:
                st.warning("No schemas available. Please create a schema first.")
                return
        
        available_schemas = schemas_df[~schemas_df.get("Archived", False)] if "Archived" in schemas_df.columns else schemas_df
        if available_schemas.empty:
            st.warning("No non-archived schemas available.")
            return
            
        schema_name = st.selectbox("Schema", available_schemas["Name"], key="admin_project_schema", help="Select the schema that defines questions for this project")
    except Exception as e:
        st.error(f"Error loading schemas: {str(e)}")
        return
    
    # Video selection with search-first approach
    st.markdown("**📹 Video Selection**")
    video_search = st.text_input("🔍 Search videos to add", placeholder="Enter video UID (min 3 chars)...", key="admin_project_video_search", help="Enter at least 3 characters to search")
    
    # Initialize selected videos in session state
    if "create_project_selected_videos" not in st.session_state:
        st.session_state.create_project_selected_videos = []
    
    # Show search results
    if len(video_search) >= 3:
        try:
            with get_db_session() as session:
                matching_videos = VideoService.search_videos_for_selection(search_term=video_search, limit=20, session=session)
            
            if matching_videos:
                if len(matching_videos) == 20:
                    st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term to find your videos faster.")
                else:
                    custom_info(f"✅ Found {len(matching_videos)} matching videos")
                
                for video in matching_videos:
                    video_col1, video_col2 = st.columns([1, 3])
                    
                    with video_col1:
                        is_selected = video['uid'] in st.session_state.create_project_selected_videos
                        if st.checkbox(f"Select", value=is_selected, key=f"video_select_{video['id']}"):
                            if video['uid'] not in st.session_state.create_project_selected_videos:
                                st.session_state.create_project_selected_videos.append(video['uid'])
                        else:
                            if video['uid'] in st.session_state.create_project_selected_videos:
                                st.session_state.create_project_selected_videos.remove(video['uid'])
                    
                    with video_col2:
                        st.markdown(f"**{video['uid']}** - {video['url'][:80]}{'...' if len(video['url']) > 80 else ''}")
            else:
                custom_info(f"No videos found matching '{video_search}'. Try a different search term.")
        except Exception as e:
            st.error(f"Error searching videos: {str(e)}")
    else:
        custom_info("Enter at least 3 characters in the search box to find videos to add")
    
    # Show selected videos
    if st.session_state.create_project_selected_videos:
        st.markdown(f"**Selected {len(st.session_state.create_project_selected_videos)} videos:**")
        
        selected_col1, selected_col2 = st.columns([1, 3])
        
        with selected_col1:
            if st.button("Clear All", key="clear_selected_videos"):
                st.session_state.create_project_selected_videos = []
                st.rerun()
        
        with selected_col2:
            for video_uid in st.session_state.create_project_selected_videos[:10]:
                st.text(f"• {video_uid}")
            if len(st.session_state.create_project_selected_videos) > 10:
                st.caption(f"... and {len(st.session_state.create_project_selected_videos) - 10} more")
    else:
        custom_info("Search and select videos above to continue")
    
    if st.button("🚀 Create Project", key="admin_create_project_btn", type="primary", use_container_width=True):
        if name and schema_name and st.session_state.create_project_selected_videos:
            try:
                with get_db_session() as session:
                    schema_id = SchemaService.get_schema_id_by_name(name=schema_name, session=session)
                    video_ids = ProjectService.get_video_ids_by_uids(video_uids=st.session_state.create_project_selected_videos, session=session)
                    ProjectService.create_project(name=name, description=description, schema_id=schema_id, video_ids=video_ids, session=session)
                
                # Clear selection after successful creation
                st.session_state.create_project_selected_videos = []
                custom_info("✅ Project created successfully!")
                # st.rerun(scope="fragment")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
        elif not name:
            st.error("❌ Project name is required")
        elif not st.session_state.create_project_selected_videos:
            st.error("❌ At least one video must be selected")

def show_project_edit_form(project_id: int):
    """Show the actual project editing form."""
    try:
        with get_db_session() as session:
            current_project = ProjectService.get_project_dict_by_id(project_id=project_id, session=session)
        
        if current_project:
            st.markdown(f"**Editing Project:** {current_project['name']}")
            
            # Basic project information
            st.markdown("### 📋 Basic Information")
            
            basic_col1, basic_col2 = st.columns(2)
            
            with basic_col1:
                st.text_input(
                    "Project ID (read-only)",
                    value=str(current_project['id']),
                    disabled=True,
                    key="admin_edit_project_id_display"
                )
                
                current_schema_name = "Unknown Schema"
                try:
                    with get_db_session() as session:
                        current_schema_name = SchemaService.get_schema_name_by_id(
                            schema_id=current_project['schema_id'], session=session
                        )
                except:
                    print(f"Error getting schema name for project {current_project['id']}")
                    pass
                
                st.text_input(
                    "Current Schema (read-only)",
                    value=current_schema_name,
                    disabled=True,
                    key="admin_edit_project_schema_display",
                    help="Schema cannot be changed after project creation"
                )
            
            with basic_col2:
                st.text_input(
                    "Project Name (read-only)",
                    value=current_project['name'],
                    disabled=True,
                    key="admin_edit_project_name",
                    help="Update the project name"
                )
                
                new_description = st.text_area(
                    "Description",
                    value=current_project.get('description') or "",
                    key="admin_edit_project_description",
                    help="Update the project description"
                )
            
            # Archive status
            archive_col1, archive_col2 = st.columns(2)
            
            with archive_col1:
                current_archived = current_project.get('is_archived', False)
                new_archived = st.checkbox(
                    "Archived",
                    value=current_archived,
                    key="admin_edit_project_archived",
                    help="Archive project to prevent new assignments"
                )
            
            with archive_col2:
                if current_archived:
                    st.warning("⚠️ This project is currently archived")
                else:
                    custom_info("✅ This project is active")
            
            # Project statistics
            st.markdown("### 📊 Project Statistics")
            
            try:
                with get_db_session() as session:
                    project_videos = VideoService.get_project_videos(
                        project_id=project_id, session=session
                    )
                    progress = ProjectService.progress(
                        project_id=project_id, session=session
                    )
                
                stats_col1, stats_col2, stats_col3, stats_col4 = st.columns(4)
                
                with stats_col1:
                    st.metric("📹 Videos", len(project_videos))
                with stats_col2:
                    st.metric("❓ Questions", progress['total_questions'])
                with stats_col3:
                    st.metric("📝 Total Answers", progress['total_answers'])
                with stats_col4:
                    st.metric("✅ GT Progress", f"{progress['completion_percentage']:.1f}%")
                
                # Video list
                if project_videos:
                    with st.expander(f"📹 View Project Videos ({len(project_videos)})", expanded=False):
                        video_data = []
                        for video in project_videos:
                            video_data.append({
                                "Video UID": video["uid"],
                                "URL": video["url"]
                            })
                        st.dataframe(pd.DataFrame(video_data), use_container_width=True)
                
            except Exception as e:
                st.error(f"Error loading project statistics: {str(e)}")
            
            # Update button
            update_col, preview_col = st.columns(2)
            
            with update_col:
                if st.button("💾 Update Project", key="admin_update_project_btn", use_container_width=True):
                    try:
                        changes_made = []
                        
                        # Note: ProjectService.update_project() needs to be implemented
                        # For now, we'll handle archive status changes
                        if new_archived != current_archived:
                            changes_made.append("Archive status")
                            with get_db_session() as session:
                                if new_archived:
                                    ProjectService.archive_project(
                                        project_id=project_id, session=session
                                    )
                                else:
                                    ProjectService.unarchive_project(
                                        project_id=project_id, session=session
                                    )
                        
                        if new_description != current_project.get('description'):
                            changes_made.append("Description")
                            with get_db_session() as session:
                                ProjectService.update_project_description(
                                    project_id=project_id, description=new_description, session=session
                                )
                        
                        if changes_made:
                            custom_info(f"✅ Project updated successfully! Changed: {', '.join(changes_made)}")
                            get_cached_project_metadata.clear()
                        else:
                            custom_info("No changes were made")
                        
                        # st.rerun(scope="fragment")
                    except Exception as e:
                        st.error(f"❌ Error updating project: {str(e)}")
            
            with preview_col:
                with st.expander("👁️ Preview Changes"):
                    st.markdown("**Changes to be applied:**")
                    # if new_name != current_project['name']:
                    #     st.markdown(f"**Name:** {current_project['name']} → {new_name}")
                    if new_description != (current_project.get('description') or ""):
                        st.markdown(f"**Description:** Updated")
                    if new_archived != current_archived:
                        status_change = "Archived" if new_archived else "Unarchived"
                        st.markdown(f"**Status:** {status_change}")
                    
                    if (new_description == (current_project.get('description') or "") and
                        new_archived == current_archived):
                        custom_info("No changes to preview")
        else:
            st.error(f"Project with ID '{project_id}' not found")
    except Exception as e:
        st.error(f"Error loading project details: {str(e)}")


@st.fragment
def admin_assignments():
    st.subheader("🔗 Project Assignments")
    
    # Load only counts for metrics - much faster
    try:
        with get_db_session() as session:
            assignment_counts = AuthService.get_assignment_counts(session=session)
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("👥 Users with Assignments", assignment_counts["unique_users"])
        with col2:
            st.metric("📁 Projects with Assignments", assignment_counts["unique_projects"])
        with col3:
            st.metric("✅ Active Assignments", assignment_counts["active_assignments"])
        with col4:
            st.metric("🗄️ Archived Assignments", assignment_counts["archived_assignments"])
            
    except Exception as e:
        st.error(f"Error loading assignment summary: {str(e)}")
        assignment_counts = {"active_assignments": 0}
    
    # Option to view assignments - now as a dialog
    if assignment_counts["active_assignments"] > 0:
        if st.button("🔍 View & Search Assignments", use_container_width=True, type="primary"):
            clear_other_dialogs('assignments')
            st.rerun()
        
        if st.session_state.get('show_assignments_dialog', False):
            st.session_state['show_assignments_dialog'] = False
            show_assignments_table_dialog()
    else:
        custom_info("No active assignments found in the database.")
    
    # Assignment management section (no expander, directly visible)
    st.markdown("### 🎯 Assign Users to Projects")
    show_assignment_management_optimized()

@st.fragment
def show_assignment_management_optimized():
    """Optimized assignment management with search-first approach and consistent styling."""
    
    # Initialize session state
    if "selected_project_ids_opt" not in st.session_state:
        st.session_state.selected_project_ids_opt = []
    if "selected_user_ids_opt" not in st.session_state:
        st.session_state.selected_user_ids_opt = []
    
    # Step 1: Project Selection with Search
    st.markdown("**Step 1: Select Projects**")
    
    project_search = st.text_input(
        "🔍 Search projects", 
        placeholder="Enter project name (min 3 chars)...", 
        key="proj_search_opt"
    )
    
    if len(project_search) >= 3:
        try:
            with get_db_session() as session:
                matching_projects = AuthService.search_projects_for_assignment(
                    search_term=project_search, limit=20, session=session
                )
            
            if matching_projects:
                if len(matching_projects) == 20:
                    st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term to find your projects faster.")
                else:
                    custom_info(f"✅ Found {len(matching_projects)} matching projects")
                
                # Batch selection buttons
                batch_col1, batch_col2 = st.columns(2)
                with batch_col1:
                    if st.button("✅ Select All Found", key="select_all_found_projects"):
                        for p in matching_projects:
                            if p['id'] not in st.session_state.selected_project_ids_opt:
                                st.session_state.selected_project_ids_opt.append(p['id'])
                        st.rerun(scope="fragment")
                
                with batch_col2:
                    if st.button("❌ Clear Selection", key="clear_selected_projects_opt"):
                        st.session_state.selected_project_ids_opt = []
                        st.rerun()
                
                # Individual project selection
                for project in matching_projects:
                    is_selected = project['id'] in st.session_state.selected_project_ids_opt
                    
                    col1, col2 = st.columns([1, 4])
                    with col1:
                        if st.checkbox("Select", value=is_selected, key=f"proj_sel_{project['id']}"):
                            if project['id'] not in st.session_state.selected_project_ids_opt:
                                st.session_state.selected_project_ids_opt.append(project['id'])
                        else:
                            if project['id'] in st.session_state.selected_project_ids_opt:
                                st.session_state.selected_project_ids_opt.remove(project['id'])
                    
                    with col2:
                        st.markdown(f"**{project['name']}** (ID: {project['id']})")
            else:
                custom_info(f"No projects found matching '{project_search}'")
        except Exception as e:
            st.error(f"Error searching projects: {str(e)}")
    
    # Show selected projects with consistent styling
    if st.session_state.selected_project_ids_opt:
        try:
            # Get project names organized by groups
            selected_by_group = {}
            for project_id in st.session_state.selected_project_ids_opt:
                try:
                    with get_db_session() as session:
                        project_info = ProjectService.get_project_dict_by_id(project_id=project_id, session=session)
                    if project_info:
                        project_name = project_info['name']
                        # Get project group if available
                        try:
                            with get_db_session() as session:
                                project_groups = ProjectGroupService.list_project_groups(session=session)
                                project_group = "Ungrouped"
                                for group in project_groups:
                                    group_info = ProjectGroupService.get_project_group_by_id(group_id=group["id"], session=session)
                                    if any(p["id"] == project_id for p in group_info["projects"]):
                                        project_group = group["name"]
                                        break
                        except:
                            project_group = "Ungrouped"
                        
                        if project_group not in selected_by_group:
                            selected_by_group[project_group] = []
                        selected_by_group[project_group].append(project_name)
                    else:
                        if "Ungrouped" not in selected_by_group:
                            selected_by_group["Ungrouped"] = []
                        selected_by_group["Ungrouped"].append(f"Project {project_id}")
                except:
                    if "Ungrouped" not in selected_by_group:
                        selected_by_group["Ungrouped"] = []
                    selected_by_group["Ungrouped"].append(f"Project {project_id}")
            
            selected_count = len(st.session_state.selected_project_ids_opt)
            if selected_by_group:
                custom_info(f"✅ <b>{selected_count}</b> projects selected")

                # Display selected projects organized by group
                for group_name, project_names in selected_by_group.items():
                    if len(selected_by_group) > 1:  # Only show group headers if multiple groups
                        custom_info(f"📁 <b>{group_name}</b> ({len(project_names)} projects): {', '.join(project_names)}")
                    else:
                        # If only one group, don't show group header
                        custom_info(f"📁 Selected {len(project_names)} projects from <b>{group_name}</b>: {', '.join(project_names)}")
            else:
                custom_info(f"✅ <b>{selected_count}</b> projects selected")
                    
        except Exception as e:
            st.error(f"Error loading project names: {str(e)}")
            custom_info(f"📁 Selected {len(st.session_state.selected_project_ids_opt)} projects")
    else:
        custom_info("Search and select projects above to continue.")
        return
    
    # Step 2: User Selection with Search
    st.markdown("**Step 2: Select Users**")
    
    user_col1, user_col2 = st.columns(2)
    with user_col1:
        user_search = st.text_input(
            "🔍 Search users", 
            placeholder="Name or email (min 3 chars)...", 
            key="user_search_opt"
        )
    with user_col2:
        user_role_filter = st.selectbox("Filter by role", ["All", "admin", "human", "model"], key="user_role_filter_opt")
    
    if len(user_search) >= 3:
        try:
            with get_db_session() as session:
                matching_users = AuthService.search_users_for_assignment(
                    search_term=user_search, 
                    user_role_filter=user_role_filter,
                    limit=20, 
                    session=session
                )
            
            if matching_users:
                if len(matching_users) == 20:
                    st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term to find your users faster.")
                else:
                    custom_info(f"✅ Found {len(matching_users)} matching users")
                
                # Batch selection buttons
                batch_user_col1, batch_user_col2 = st.columns(2)
                with batch_user_col1:
                    if st.button("✅ Select All Found", key="select_all_found_users"):
                        for u in matching_users:
                            if u['id'] not in st.session_state.selected_user_ids_opt:
                                st.session_state.selected_user_ids_opt.append(u['id'])
                        st.rerun(scope="fragment")
                
                with batch_user_col2:
                    if st.button("❌ Clear Selection", key="clear_selected_users_opt"):
                        st.session_state.selected_user_ids_opt = []
                        st.rerun(scope="fragment")
                
                # Individual user selection
                for user in matching_users:
                    is_selected = user['id'] in st.session_state.selected_user_ids_opt
                    
                    col1, col2 = st.columns([1, 4])
                    with col1:
                        if st.checkbox("Select", value=is_selected, key=f"user_sel_{user['id']}"):
                            if user['id'] not in st.session_state.selected_user_ids_opt:
                                st.session_state.selected_user_ids_opt.append(user['id'])
                        else:
                            if user['id'] in st.session_state.selected_user_ids_opt:
                                st.session_state.selected_user_ids_opt.remove(user['id'])
                    
                    with col2:
                        role_emoji = {"human": "👤", "admin": "👑", "model": "🤖"}.get(user['role'], "❓")
                        if user['role'] == 'model':
                            st.markdown(f"{role_emoji} **{user['name']}** - {user['role']}")
                        else:
                            st.markdown(f"{role_emoji} **{user['name']}** ({user['email']}) - {user['role']}")
            else:
                custom_info(f"No users found matching '{user_search}' with role filter '{user_role_filter}'")
        except Exception as e:
            st.error(f"Error searching users: {str(e)}")
    
    # Show selected users with consistent styling
    if st.session_state.selected_user_ids_opt:
        try:
            # Get user names organized by role
            selected_by_role = {}
            for user_id in st.session_state.selected_user_ids_opt:
                try:
                    with get_db_session() as session:
                        user_info = AuthService.get_user_info_by_id(user_id=user_id, session=session)
                    if user_info:
                        user_role = user_info['user_type']
                        role_emoji = {"human": "👤", "admin": "👑", "model": "🤖"}.get(user_role, "❓")
                        
                        if user_role not in selected_by_role:
                            selected_by_role[user_role] = []
                        
                        if user_role == 'model':
                            selected_by_role[user_role].append(f"{role_emoji} {user_info['user_id_str']}")
                        else:
                            selected_by_role[user_role].append(f"{role_emoji} {user_info['user_id_str']} ({user_info['email']})")
                    else:
                        if "unknown" not in selected_by_role:
                            selected_by_role["unknown"] = []
                        selected_by_role["unknown"].append(f"❓ User {user_id}")
                except:
                    if "unknown" not in selected_by_role:
                        selected_by_role["unknown"] = []
                    selected_by_role["unknown"].append(f"❓ User {user_id}")
            
            selected_count = len(st.session_state.selected_user_ids_opt)
            if selected_by_role:
                custom_info(f"✅ <b>{selected_count}</b> users selected")

                # Display selected users organized by role
                for role_name, user_names in selected_by_role.items():
                    role_display = {"human": "Human Users", "admin": "Admin Users", "model": "Model Users"}.get(role_name, "Unknown Users")
                    if len(selected_by_role) > 1:  # Only show role headers if multiple roles
                        custom_info(f"👥 <b>{role_display}</b> ({len(user_names)} users): {', '.join(user_names)}")
                    else:
                        # If only one role, don't show role header
                        custom_info(f"👥 Selected {len(user_names)} <b>{role_display.lower()}</b>: {', '.join(user_names)}")
            else:
                custom_info(f"✅ <b>{selected_count}</b> users selected")
                    
        except Exception as e:
            st.error(f"Error loading user names: {str(e)}")
            custom_info(f"👥 Selected {len(st.session_state.selected_user_ids_opt)} users")
    else:
        custom_info("Search and select users above to continue.")
        return
    
    # Step 3: Assignment Settings
    st.markdown("**Step 3: Assignment Settings**")
    
    settings_col1, settings_col2 = st.columns(2)
    
    with settings_col1:
        role = st.selectbox("Assignment Role", ["annotator", "reviewer", "admin", "model"], key="assign_role_opt")
    
    with settings_col2:
        user_weight = st.number_input(
            "User Weight", 
            min_value=0.0, 
            value=1.0, 
            step=0.1,
            key="assign_user_weight_opt",
            help="Weight for user's answers in scoring"
        )
    
    # Show assignment preview
    total_operations = len(st.session_state.selected_project_ids_opt) * len(st.session_state.selected_user_ids_opt)
    custom_info(f"🎯 Ready to create {total_operations} assignments ({len(st.session_state.selected_user_ids_opt)} users × {len(st.session_state.selected_project_ids_opt)} projects)")
    
    # Assignment actions
    action_col1, action_col2 = st.columns(2)
    
    with action_col1:
        if st.button("✅ Execute Assignments", key="execute_assignments_opt", use_container_width=True):
            try:
                with get_db_session() as session:
                    # First verify ALL assignments would succeed
                    AuthService.verify_bulk_assign_users_to_projects(
                        user_ids=st.session_state.selected_user_ids_opt,
                        project_ids=st.session_state.selected_project_ids_opt,
                        role=role,
                        session=session
                    )
                    
                    # If verification passes, execute ALL assignments
                    total_assignments = AuthService.bulk_assign_users_to_projects(
                        user_ids=st.session_state.selected_user_ids_opt,
                        project_ids=st.session_state.selected_project_ids_opt,
                        role=role,
                        session=session,
                        user_weight=user_weight
                    )
                
                custom_info(f"✅ Successfully completed {total_assignments} assignments!")
                # Clear selections after success
                st.session_state.selected_project_ids_opt = []
                st.session_state.selected_user_ids_opt = []
                
            except ValueError as e:
                st.error(f"❌ Assignment failed: {str(e)}")
            except Exception as e:
                st.error(f"❌ Unexpected error: {str(e)}")
 

    with action_col2:
        if st.button("🗑️ Remove Assignments", key="execute_removals_opt", use_container_width=True):
            try:
                with get_db_session() as session:
                    # First verify ALL removals would succeed
                    AuthService.verify_bulk_remove_users_from_projects(
                        user_ids=st.session_state.selected_user_ids_opt,
                        project_ids=st.session_state.selected_project_ids_opt,
                        role=role,
                        session=session
                    )
                
                    # If verification passes, execute ALL removals
                    total_removals = AuthService.bulk_remove_users_from_projects(
                        user_ids=st.session_state.selected_user_ids_opt,
                        project_ids=st.session_state.selected_project_ids_opt,
                        role=role,
                        session=session
                    )
                
                custom_info(f"🗑️ Successfully removed {total_removals} assignments!")
                # Clear selections after success
                st.session_state.selected_project_ids_opt = []
                st.session_state.selected_user_ids_opt = []
                
            except ValueError as e:
                st.error(f"❌ Removal failed: {str(e)}")
            except Exception as e:
                st.error(f"❌ Unexpected error: {str(e)}")


@st.fragment
def admin_schemas():
    st.subheader("📋 Schema Management")
    
    # Load only counts for metrics - much faster
    try:
        with get_db_session() as session:
            schema_counts = SchemaService.get_schema_counts(session=session)
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📋 Total Schemas", schema_counts["total"])
        with col2:
            st.metric("✅ Active Schemas", schema_counts["active"])
        with col3:
            st.metric("🗄️ Archived Schemas", schema_counts["archived"])
        with col4:
            st.metric("🎨 Custom Display", schema_counts["custom_display"])
    except Exception as e:
        st.error(f"Error loading schema summary: {str(e)}")
        schema_counts = {"total": 0}
    
    # Option to view schemas table - now as a dialog
    if schema_counts["total"] > 0:
        if st.button("📋 View All Schemas (Database Table)", use_container_width=True):
            clear_other_dialogs('schemas')
            st.rerun()
        
        if st.session_state.get('show_schemas_dialog', False):
            st.session_state['show_schemas_dialog'] = False
            show_schemas_table_dialog()
    
    # Management section
    schema_management_tabs = st.tabs(["➕ Create Schema", "✏️ Edit Schema"])
    
    with schema_management_tabs[0]:
        show_schema_create_form()
    
    with schema_management_tabs[1]:
        show_schema_edit_interface()

def show_schema_create_form():
    """Show schema creation form with search-optimized question group selection."""
    st.markdown("### 🆕 Create New Schema")
    
    schema_name = st.text_input("Schema Name", key="admin_schema_name", placeholder="Enter schema name...")
    
    # Instructions URL
    instructions_url = st.text_input("Instructions URL (optional)", key="admin_schema_instructions", 
                                   placeholder="https://example.com/instructions", 
                                   help="URL linking to instructions for this schema")
    
    # Custom display checkbox
    has_custom_display = st.checkbox("Has Custom Display", key="admin_schema_custom_display", 
                                   help="Enable if this schema has custom display logic for questions or options")
    
    # Question groups selection with search-first approach
    st.markdown("**📁 Select Question Groups:**")
    
    # Check if we have many groups to decide on approach
    try:
        with get_db_session() as session:
            groups_count = QuestionGroupService.get_group_counts(session=session)
        total_groups = groups_count.get("total", 0)
            
        if total_groups > 20:
            # Use search-first approach for many groups
            group_search = st.text_input("🔍 Search question groups to add", placeholder="Enter group name (min 3 chars)...", key="admin_schema_group_search", help="Enter at least 3 characters to search")
            
            # Initialize selected groups in session state - NOW STORE OBJECTS WITH ID AND NAME
            if "create_schema_selected_groups" not in st.session_state:
                st.session_state.create_schema_selected_groups = []  # List of {"id": int, "name": str, "display_title": str}
            
            # Show search results
            if len(group_search) >= 3:
                try:
                    with get_db_session() as session:
                        matching_groups = QuestionGroupService.search_groups_for_selection(search_term=group_search, limit=20, session=session)
                    
                    if matching_groups:
                        if len(matching_groups) == 20:
                            st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term to find your groups faster.")
                        else:
                            custom_info(f"✅ Found {len(matching_groups)} matching groups")
                        
                        for group in matching_groups:
                            if not group.get('archived', False):  # Only show non-archived
                                group_col1, group_col2 = st.columns([1, 3])
                                
                                with group_col1:
                                    # Check if already selected using ID
                                    selected_ids = [g["id"] for g in st.session_state.create_schema_selected_groups]
                                    is_selected = group['id'] in selected_ids
                                    
                                    if st.checkbox(f"Select", value=is_selected, key=f"group_select_{group['id']}"):
                                        if group['id'] not in selected_ids:
                                            # Store full object with ID and names
                                            st.session_state.create_schema_selected_groups.append({
                                                "id": group['id'],
                                                "name": group['title'], 
                                                "display_title": group['display_title']
                                            })
                                    else:
                                        # Remove from selection
                                        st.session_state.create_schema_selected_groups = [
                                            g for g in st.session_state.create_schema_selected_groups 
                                            if g["id"] != group['id']
                                        ]
                                
                                with group_col2:
                                    description_preview = group.get('description', '')[:50] + '...' if len(str(group.get('description', ''))) > 50 else group.get('description', '')
                                    st.markdown(f"**{group['display_title']}** (Internal: {group['title']})")
                                    if description_preview:
                                        st.caption(description_preview)
                    else:
                        custom_info(f"No question groups found matching '{group_search}'. Try a different search term.")
                except Exception as e:
                    st.error(f"Error searching question groups: {str(e)}")
            else:
                custom_info("Enter at least 3 characters in the search box to find question groups")
            
            # Show selected groups - NOW WITH PROPER NAMES
            if st.session_state.create_schema_selected_groups:
                st.markdown(f"**Selected {len(st.session_state.create_schema_selected_groups)} groups (in order):**")
                
                selected_col1, selected_col2 = st.columns([1, 3])
                
                with selected_col1:
                    if st.button("Clear All", key="clear_selected_groups_schema"):
                        st.session_state.create_schema_selected_groups = []
                        st.rerun()
                
                with selected_col2:
                    for i, group_obj in enumerate(st.session_state.create_schema_selected_groups[:10]):
                        # NOW SHOW ACTUAL NAMES!
                        st.text(f"{i+1}. {group_obj['display_title']}")
                    if len(st.session_state.create_schema_selected_groups) > 10:
                        st.caption(f"... and {len(st.session_state.create_schema_selected_groups) - 10} more")
            
            # Extract just IDs for the service call
            selected_groups = [g["id"] for g in st.session_state.create_schema_selected_groups]
            
        else:
            # Use traditional multiselect for fewer groups
            with get_db_session() as session:
                groups_df = QuestionGroupService.get_all_groups(session=session)
            if not groups_df.empty:
                available_groups = groups_df[~groups_df["Archived"]]
                if not available_groups.empty:
                    selected_groups = st.multiselect(
                        "Question Groups (in desired order)", 
                        available_groups["ID"].tolist(),
                        format_func=lambda x: f"{available_groups[available_groups['ID']==x]['Name'].iloc[0]} - {available_groups[available_groups['ID']==x]['Description'].iloc[0][:50]}{'...' if len(str(available_groups[available_groups['ID']==x]['Description'].iloc[0])) > 50 else ''}",
                        key="admin_schema_groups",
                        help="Select question groups in the order they should appear"
                    )
                    
                    if selected_groups:
                        st.markdown("**📋 Selected Groups (Preview Order):**")
                        for i, group_id in enumerate(selected_groups):
                            group_name = available_groups[available_groups["ID"]==group_id]["Name"].iloc[0]
                            st.markdown(f"{i+1}. {group_name}")
                else:
                    st.warning("No non-archived question groups available.")
                    selected_groups = []
            else:
                st.warning("No question groups available.")
                selected_groups = []
    
    except Exception as e:
        st.error(f"Error loading question groups: {str(e)}")
        selected_groups = []
    
    if st.button("🚀 Create Schema", key="admin_create_schema_btn", type="primary", use_container_width=True):
        if schema_name and selected_groups:
            try:
                # Clean up instructions URL
                clean_instructions_url = instructions_url.strip() if instructions_url else None
                if clean_instructions_url == "":
                    clean_instructions_url = None
                
                with get_db_session() as session:
                    SchemaService.create_schema(
                        name=schema_name, 
                        question_group_ids=selected_groups, 
                        instructions_url=clean_instructions_url,
                        has_custom_display=has_custom_display,
                        session=session
                    )
                
                # Clear selection after successful creation
                if "create_schema_selected_groups" in st.session_state:
                    st.session_state.create_schema_selected_groups = []
                
                custom_info("✅ Schema created successfully!")
                # st.rerun(scope="fragment")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
        elif not schema_name:
            st.error("❌ Schema name is required")
        elif not selected_groups:
            st.error("❌ At least one question group must be selected")


def show_schema_edit_interface():
    """Show schema editing interface with search-first approach."""
    st.markdown("### ✏️ Edit Schema")
    
    # Search-first approach
    search_term = st.text_input(
        "🔍 Search for schema to edit", 
        placeholder="Enter schema name to search...",
        key="admin_edit_schema_search",
        help="Enter at least 3 characters to search"
    )
    
    if len(search_term) >= 3:
        try:
            with get_db_session() as session:
                # Search for schemas matching the term
                matching_schemas = SchemaService.search_schemas_for_selection(
                    search_term=search_term, limit=20, session=session
                )
            
            if matching_schemas:
                if len(matching_schemas) == 20:
                    st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term.")
                else:
                    custom_info(f"✅ Found {len(matching_schemas)} matching schemas")
                
                # Show matching schemas for selection
                schema_options = {}
                for s in matching_schemas:
                    archived_indicator = " 🗄️ (ARCHIVED)" if s['archived'] else ""
                    schema_options[f"{s['name']}{archived_indicator} (ID: {s['id']})"] = s['id']
                
                selected_schema_display = st.selectbox(
                    f"Select from {len(matching_schemas)} matching schemas",
                    [""] + list(schema_options.keys()),
                    key="admin_edit_schema_select"
                )
                
                if selected_schema_display:
                    selected_schema_id = schema_options[selected_schema_display]
                    show_schema_edit_form(selected_schema_id)
            else:
                custom_info(f"No schemas found matching '{search_term}'")
        except Exception as e:
            st.error(f"Error searching schemas: {str(e)}")
    else:
        custom_info("Enter at least 3 characters in the search box to find schemas to edit")


def show_schema_edit_form(schema_id: int):
    """Show the actual schema editing form."""
    try:
        with get_db_session() as session:
            schema_details = SchemaService.get_schema_details(schema_id=schema_id, session=session)
        
        st.markdown(f"**Editing Schema:** {schema_details['name']}")
        
        # Basic schema information editing
        st.markdown("### 📋 Basic Information")
        
        new_name = st.text_input(
            "Schema Name",
            value=schema_details["name"],
            key="admin_edit_schema_name"
        )
        
        new_instructions_url = st.text_input(
            "Instructions URL",
            value=schema_details["instructions_url"] or "",
            key="admin_edit_schema_instructions",
            placeholder="https://example.com/instructions",
            help="Leave empty to remove instructions URL"
        )
        
        edit_col1, edit_col2 = st.columns(2)
        with edit_col1:
            new_has_custom_display = st.checkbox(
                "Has Custom Display",
                value=schema_details["has_custom_display"],
                key="admin_edit_schema_custom_display",
                help="Enable if this schema uses custom display logic for questions or options"
            )
        
        with edit_col2:
            new_is_archived = st.checkbox(
                "Archived",
                value=schema_details["is_archived"],
                key="admin_edit_schema_archived",
                help="Archive schema to prevent use in new projects"
            )
        
        # Question group order management
        st.markdown("### 📁 Question Group Order Management")
        with get_db_session() as session:
            current_order = SchemaService.get_question_group_order(schema_id=schema_id, session=session)
        
        if current_order:
            with get_db_session() as session:
                groups_df = QuestionGroupService.get_all_groups(session=session)
            
            order_key = f"edit_schema_order_{schema_id}"
            if order_key not in st.session_state:
                st.session_state[order_key] = current_order.copy()
            
            working_order = st.session_state[order_key]
            
            custom_info("💡 Use the ⬆️ and ⬇️ buttons to reorder question groups. Changes will be applied when you click 'Update Schema'.")
            
            if len(working_order) > 5:
                search_term = st.text_input(
                    "🔍 Search question groups (to quickly find groups in large schemas)",
                    key=f"search_groups_{schema_id}",
                    placeholder="Type part of a group name..."
                )
            else:
                search_term = ""
            
            for i, group_id in enumerate(working_order):
                group_row = groups_df[groups_df["ID"] == group_id]
                if not group_row.empty:
                    group_name = group_row.iloc[0]["Name"]
                    group_display_title = group_row.iloc[0]["Display Title"]
                    group_description = group_row.iloc[0]["Description"]
                    is_archived_group = group_row.iloc[0]["Archived"]
                    
                    if search_term and search_term.lower() not in group_name.lower() and search_term.lower() not in group_display_title.lower():
                        continue
                    
                    group_order_col1, group_order_col2, group_order_col3 = st.columns([0.1, 0.8, 0.1])
                    
                    with group_order_col1:
                        if st.button("⬆️", key=f"group_up_{schema_id}_{group_id}_{i}", 
                                    disabled=(i == 0), help="Move up"):
                            st.session_state[order_key][i], st.session_state[order_key][i-1] = \
                                st.session_state[order_key][i-1], st.session_state[order_key][i]
                            st.rerun()
                    
                    with group_order_col2:
                        status_icon = "🗄️" if is_archived_group else "✅"
                        display_text = f"{group_display_title} (Internal: {group_name})"
                        if len(display_text) > 80:
                            display_text = display_text[:80] + '...'
                        
                        if search_term and (search_term.lower() in group_name.lower() or search_term.lower() in group_display_title.lower()):
                            st.write(f"**{i+1}.** {status_icon} {display_text} 🔍")
                        else:
                            st.write(f"**{i+1}.** {status_icon} {display_text}")
                        
                        if group_description:
                            st.caption(f"Description: {group_description[:100]}{'...' if len(group_description) > 100 else ''}")
                        st.caption(f"ID: {group_id}")
                    
                    with group_order_col3:
                        if st.button("⬇️", key=f"group_down_{schema_id}_{group_id}_{i}", 
                                    disabled=(i == len(working_order) - 1), help="Move down"):
                            st.session_state[order_key][i], st.session_state[order_key][i+1] = \
                                st.session_state[order_key][i+1], st.session_state[order_key][i]
                            st.rerun()
            
            group_order_action_col1, group_order_action_col2 = st.columns(2)
            with group_order_action_col1:
                if st.button("🔄 Reset Group Order", key=f"reset_group_order_{schema_id}"):
                    st.session_state[order_key] = current_order.copy()
                    st.rerun()
            
            with group_order_action_col2:
                if working_order != current_order:
                    st.warning("⚠️ Group order changed - click 'Update Schema' to save")
                else:
                    custom_info("✅ Order matches saved state")
            
            new_group_order = working_order
        else:
            new_group_order = current_order
            custom_info("No question groups in this schema.")
        
        # Update button with preview
        update_col, preview_col = st.columns(2)
        
        with update_col:
            if st.button("💾 Update Schema", key="admin_update_schema_btn", type="primary", use_container_width=True):
                try:
                    changes_made = []
                    
                    # Check what changed
                    if new_name != schema_details["name"]:
                        changes_made.append("Name")
                    
                    clean_instructions_url = new_instructions_url.strip() if new_instructions_url else ""
                    current_instructions = schema_details["instructions_url"] or ""
                    if clean_instructions_url != current_instructions:
                        changes_made.append("Instructions URL")
                    
                    if new_has_custom_display != schema_details["has_custom_display"]:
                        changes_made.append("Custom display setting")
                    
                    if new_is_archived != schema_details["is_archived"]:
                        changes_made.append("Archive status")
                    
                    if new_group_order != current_order:
                        changes_made.append("Question group order")
                    
                    if changes_made:
                        # Prepare instructions URL (empty string means clear it)
                        if clean_instructions_url == "":
                            clean_instructions_url = ""  # This will be handled by edit_schema to set to None
                        
                        with get_db_session() as session:
                            SchemaService.edit_schema(
                                schema_id=schema_id,
                                name=new_name if new_name != schema_details["name"] else None,
                                instructions_url=clean_instructions_url if clean_instructions_url != current_instructions else None,
                                has_custom_display=new_has_custom_display if new_has_custom_display != schema_details["has_custom_display"] else None,
                                is_archived=new_is_archived if new_is_archived != schema_details["is_archived"] else None,
                                session=session
                            )
                        
                        # Update question group order if changed
                        if new_group_order != current_order:
                            with get_db_session() as session:
                                SchemaService.update_question_group_order(
                                    schema_id=schema_id, 
                                    group_ids=new_group_order, 
                                    session=session
                                )
                        
                        # Clear order state after successful update
                        order_key = f"edit_schema_order_{schema_id}"
                        if order_key in st.session_state:
                            del st.session_state[order_key]
                        
                        custom_info(f"✅ Schema updated successfully! Changed: {', '.join(changes_made)}")
                        # st.rerun(scope="fragment")
                    else:
                        custom_info("No changes were made")
                        
                except Exception as e:
                    st.error(f"❌ Error updating schema: {str(e)}")
        
        with preview_col:
            with st.expander("👁️ Preview Changes"):
                st.markdown("**Changes to be applied:**")
                
                if new_name != schema_details["name"]:
                    st.markdown(f"**Name:** {schema_details['name']} → {new_name}")
                
                clean_instructions_url = new_instructions_url.strip() if new_instructions_url else ""
                current_instructions = schema_details["instructions_url"] or ""
                if clean_instructions_url != current_instructions:
                    old_display = current_instructions or "None"
                    new_display = clean_instructions_url or "None"
                    st.markdown(f"**Instructions URL:** {old_display} → {new_display}")
                
                if new_has_custom_display != schema_details["has_custom_display"]:
                    st.markdown(f"**Custom Display:** {schema_details['has_custom_display']} → {new_has_custom_display}")
                
                if new_is_archived != schema_details["is_archived"]:
                    status_change = "Archived" if new_is_archived else "Unarchived"
                    st.markdown(f"**Status:** {status_change}")
                
                if new_group_order != current_order:
                    st.markdown(f"**Group Order:** Reordered")
                
                # Check if no changes
                no_changes = (
                    new_name == schema_details["name"] and
                    clean_instructions_url == current_instructions and
                    new_has_custom_display == schema_details["has_custom_display"] and
                    new_is_archived == schema_details["is_archived"] and
                    new_group_order == current_order
                )
                
                if no_changes:
                    custom_info("No changes to preview")
                    
    except Exception as e:
        st.error(f"Error loading schema details: {str(e)}")


@st.fragment
def admin_questions():
    st.subheader("❓ Question & Group Management")
    
    q_tab1, q_tab2 = st.tabs(["📁 Question Groups", "❓ Individual Questions"])
    
    with q_tab1:
        # Get counts only for metrics - much faster
        try:
            with get_db_session() as session:
                group_counts = QuestionGroupService.get_group_counts(session=session)
            
            # Summary metrics
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("📁 Total Groups", group_counts["total"])
            with col2:
                st.metric("✅ Active Groups", group_counts["active"])
            with col3:
                st.metric("🗄️ Archived Groups", group_counts["archived"])
            with col4:
                st.metric("🔄 Reusable Groups", group_counts["reusable"])
        except Exception as e:
            st.error(f"Error loading group summary: {str(e)}")
            group_counts = {"total": 0}
        
        # Option to view groups table - now as a dialog
        if group_counts["total"] > 0:
            if st.button("📋 View All Question Groups (Database Table)", use_container_width=True):
                clear_other_dialogs('question_groups')
                st.rerun()
            
            if st.session_state.get('show_question_groups_dialog', False):
                st.session_state['show_question_groups_dialog'] = False
                show_question_groups_table_dialog()
        
        # Management tabs for groups
        group_management_tabs = st.tabs(["➕ Create Group", "✏️ Edit Group"])
        
        with group_management_tabs[0]:
            st.markdown("### 🆕 Create New Question Group")
            
            title = st.text_input("Group Title", key="admin_group_title", placeholder="Enter group title...")
            basic_col1, basic_col2 = st.columns(2)
            with basic_col1:
                is_reusable = st.checkbox("Reusable across schemas", key="admin_group_reusable", 
                                        help="Allow this group to be used in multiple schemas")
            with basic_col2:
                is_auto_submit = st.checkbox("Auto Submit", key="admin_group_auto_submit", 
                                        help="Automatically submit answers for this group")
            description = st.text_area("Description", key="admin_group_description", 
                                        placeholder="Describe the purpose of this question group...")
            
            # Verification function selection
            st.markdown("**🔧 Verification Function (Optional):**")
            try:
                available_functions = QuestionGroupService.get_available_verification_functions()
                if available_functions:
                    verification_function = st.selectbox(
                        "Select verification function",
                        ["None"] + available_functions,
                        key="admin_group_verification",
                        help="Optional function to validate answers"
                    )
                    
                    if verification_function != "None":
                        try:
                            func_info = QuestionGroupService.get_verification_function_info(verification_function)
                            custom_info(f"Function: `{func_info['name']}{func_info['signature']}`")
                            if func_info['docstring']:
                                st.markdown(f"**Documentation:** {func_info['docstring']}")
                        except Exception as e:
                            st.error(f"Error loading function info: {str(e)}")
                    
                    verification_function = verification_function if verification_function != "None" else None
                else:
                    custom_info("No verification functions found in verify.py")
                    verification_function = None
            except Exception as e:
                st.error(f"Error loading verification functions: {str(e)}")
                verification_function = None
            
            st.markdown("**📋 Select Questions:**")
            
            # Check if we have many questions to decide on approach
            try:
                with get_db_session() as session:
                    question_counts = QuestionService.get_question_counts(session=session)
                total_questions = question_counts.get("active", 0)
                
                if total_questions > 20:
                    # Use search-first approach for many questions
                    question_search = st.text_input("🔍 Search questions to add", placeholder="Enter question text (min 3 chars)...", key="admin_group_question_search", help="Enter at least 3 characters to search")
                    
                    # Initialize selected questions in session state - STORE OBJECTS WITH ID AND TEXT
                    if "create_group_selected_questions" not in st.session_state:
                        st.session_state.create_group_selected_questions = []  # List of {"id": int, "text": str}
                    
                    # Show search results
                    if len(question_search) >= 3:
                        try:
                            with get_db_session() as session:
                                matching_questions = QuestionService.search_questions_for_selection(search_term=question_search, limit=20, session=session)
                            
                            if matching_questions:
                                if len(matching_questions) == 20:
                                    st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term to find your questions faster.")
                                else:
                                    custom_info(f"✅ Found {len(matching_questions)} matching questions")
                                
                                for question in matching_questions:
                                    if not question.get('archived', False):  # Only show non-archived
                                        question_col1, question_col2 = st.columns([1, 3])
                                        
                                        with question_col1:
                                            # Check if already selected using ID
                                            selected_ids = [q["id"] for q in st.session_state.create_group_selected_questions]
                                            is_selected = question['id'] in selected_ids
                                            
                                            if st.checkbox(f"Select", value=is_selected, key=f"question_select_{question['id']}"):
                                                if question['id'] not in selected_ids:
                                                    # Store full object with ID and text
                                                    st.session_state.create_group_selected_questions.append({
                                                        "id": question['id'],
                                                        "text": question['text']
                                                    })
                                            else:
                                                # Remove from selection
                                                st.session_state.create_group_selected_questions = [
                                                    q for q in st.session_state.create_group_selected_questions 
                                                    if q["id"] != question['id']
                                                ]
                                        
                                        with question_col2:
                                            # Show question text preview
                                            text_preview = question['text'][:80] + '...' if len(question['text']) > 80 else question['text']
                                            question_type = question.get('type', 'unknown')
                                            type_emoji = "🔘" if question_type == "single" else "📝"
                                            st.markdown(f"{type_emoji} **{text_preview}**")
                                            st.caption(f"Type: {question_type}")
                            else:
                                custom_info(f"No questions found matching '{question_search}'. Try a different search term.")
                        except Exception as e:
                            st.error(f"Error searching questions: {str(e)}")
                    else:
                        custom_info("Enter at least 3 characters in the search box to find questions")
                    
                    # Show selected questions - WITH PROPER TEXT
                    if st.session_state.create_group_selected_questions:
                        st.markdown(f"**Selected {len(st.session_state.create_group_selected_questions)} questions (in order):**")
                        
                        selected_col1, selected_col2 = st.columns([1, 3])
                        
                        with selected_col1:
                            if st.button("Clear All", key="clear_selected_questions_group"):
                                st.session_state.create_group_selected_questions = []
                                st.rerun()
                        
                        with selected_col2:
                            for i, question_obj in enumerate(st.session_state.create_group_selected_questions[:10]):
                                # NOW SHOW ACTUAL QUESTION TEXT!
                                text_preview = question_obj['text'][:60] + '...' if len(question_obj['text']) > 60 else question_obj['text']
                                st.text(f"{i+1}. {text_preview}")
                            if len(st.session_state.create_group_selected_questions) > 10:
                                st.caption(f"... and {len(st.session_state.create_group_selected_questions) - 10} more")
                    
                    # Extract just IDs for the service call
                    selected_questions = [q["id"] for q in st.session_state.create_group_selected_questions]
                    
                else:
                    # Use traditional multiselect for fewer questions
                    try:
                        with get_db_session() as session:
                            questions_df_for_create = QuestionService.get_all_questions(session=session)
                        if not questions_df_for_create.empty:
                            available_questions = questions_df_for_create[~questions_df_for_create["Archived"]]
                            selected_questions = st.multiselect(
                                "Questions",
                                available_questions["ID"].tolist(),
                                format_func=lambda x: available_questions[available_questions["ID"]==x]["Text"].iloc[0][:80] + ('...' if len(available_questions[available_questions["ID"]==x]["Text"].iloc[0]) > 80 else ''),
                                key="admin_group_questions",
                                help="Select questions to include in this group"
                            )
                        else:
                            selected_questions = []
                            custom_info("No questions available.")
                    except Exception as e:
                        st.error(f"Error loading questions: {str(e)}")
                        selected_questions = []
            
            except Exception as e:
                st.error(f"Error loading question counts: {str(e)}")
                selected_questions = []
            
            if st.button("🚀 Create Question Group", key="admin_create_group_btn", type="primary", use_container_width=True):
                if title and selected_questions:
                    try:
                        with get_db_session() as session:
                            QuestionGroupService.create_group(
                                title=title, display_title=title, description=description, is_reusable=is_reusable, 
                                question_ids=selected_questions, verification_function=verification_function, 
                                is_auto_submit=is_auto_submit, session=session
                            )
                        # Clear selection after successful creation
                        if "create_group_selected_questions" in st.session_state:
                            st.session_state.create_group_selected_questions = []

                        custom_info("✅ Question group created successfully!")
                        # st.rerun(scope="fragment")
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
                elif not title:
                    st.error("❌ Group title is required")
                elif not selected_questions:
                    st.error("❌ At least one question must be selected")
        
        with group_management_tabs[1]:
            show_group_edit_interface()
    
    with q_tab2:
        # Get counts only for metrics - much faster
        try:
            with get_db_session() as session:
                question_counts = QuestionService.get_question_counts(session=session)
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("❓ Total Questions", question_counts["total"])
            with col2:
                st.metric("✅ Active Questions", question_counts["active"])
            with col3:
                st.metric("🔘 Single Choice", question_counts["single_choice"])
            with col4:
                st.metric("📝 Description", question_counts["description"])
        except Exception as e:
            st.error(f"Error loading question summary: {str(e)}")
            question_counts = {"total": 0}
        
        # Option to view questions table
        if question_counts["total"] > 0:
            if st.button("📋 View All Questions (Database Table)", use_container_width=True):
                clear_other_dialogs('questions')
                st.rerun()
        
            if st.session_state.get('show_questions_dialog', False):
                st.session_state['show_questions_dialog'] = False
                show_questions_table_dialog()
        
        # Management tabs remain the same
        question_management_tabs = st.tabs(["➕ Create Question", "✏️ Edit Question"])
        
        with question_management_tabs[0]:
            st.markdown("### 🆕 Create New Question")
            
            basic_info_col1, basic_info_col2 = st.columns(2)
            with basic_info_col1:
                text = st.text_input("Question Text", key="admin_question_text", 
                                    placeholder="Enter the question text...")
            with basic_info_col2:
                q_type = st.selectbox("Question Type", ["single", "description"], key="admin_question_type",
                                    help="Single: Multiple choice | Description: Text input")
            
            use_text_as_display = st.checkbox("Use question text as display text", value=True, 
                                            key="admin_question_use_text_as_display",
                                            help="Uncheck to provide custom display text")
            if not use_text_as_display:
                display_text = st.text_input("Question to display to user", key="admin_question_display_text", 
                                            value=text, placeholder="Text shown to users...")
            else:
                display_text = None
            
            options = []
            option_weights = []
            default = None
            
            if q_type == "single":
                st.markdown("**🎯 Options and Weights:**")
                custom_info("💡 Default weight is 1.0 for each option. Customize weights to influence scoring.")
                
                num_options = st.number_input("Number of options", 1, 10, 2, key="admin_question_num_options")
                
                for i in range(num_options):
                    opt_col1, opt_col2 = st.columns([3, 1])
                    with opt_col1:
                        option = st.text_input(f"Option {i+1}", key=f"admin_question_opt_{i}",
                                                placeholder=f"Enter option {i+1}...")
                    with opt_col2:
                        weight = st.number_input(f"Weight {i+1}", min_value=0.0, value=1.0, step=0.1, 
                                                key=f"admin_question_weight_{i}", 
                                                help="Weight for scoring (default: 1.0)")
                    
                    if option:
                        options.append(option)
                        option_weights.append(weight)
                
                if options:
                    default = st.selectbox("Default option", [""] + options, key="admin_question_default",
                                            help="Option selected by default")
                    if default == "":
                        default = None
            elif q_type == "description":
                # For description questions, allow setting a default response
                default = st.text_input("Default response (optional)", key="admin_question_default_desc",
                                        placeholder="Default text response...", 
                                        help="Optional default response for description questions")
                if not default.strip():
                    default = None
            
            if st.button("🚀 Create Question", key="admin_create_question_btn", type="primary", use_container_width=True):
                if text:
                    try:
                        with get_db_session() as session:
                            QuestionService.add_question(
                                text=text, qtype=q_type, options=options if q_type == "single" else None,
                                default=default, session=session,
                                display_text=display_text, option_weights=option_weights if q_type == "single" else None
                            )
                        custom_info("✅ Question created successfully!")
                        # st.rerun(scope="fragment")
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
        
        with question_management_tabs[1]:
            show_question_edit_interface()


def show_group_edit_interface():
    """Show group editing interface with search-first approach."""
    st.markdown("### ✏️ Edit Question Group")
    
    # Search-first approach
    search_term = st.text_input(
        "🔍 Search for group to edit", 
        placeholder="Enter group name to search...",
        key="admin_edit_group_search",
        help="Enter at least 3 characters to search"
    )
    
    if len(search_term) >= 3:
        try:
            # Search for groups matching the term
            with get_db_session() as session:
                matching_groups = QuestionGroupService.search_groups_for_selection(
                    search_term=search_term, limit=20, session=session
                )
            
            if matching_groups:
                if len(matching_groups) == 20:
                    st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term.")
                else:
                    custom_info(f"✅ Found {len(matching_groups)} matching groups")
                
                # Show matching groups for selection
                group_options = {}
                for g in matching_groups:
                    archived_indicator = " 🗄️ (ARCHIVED)" if g['archived'] else ""
                    group_options[f"{g['display_title']}{archived_indicator} (Internal: {g['title']}, ID: {g['id']})"] = g['id']
                
                selected_group_display = st.selectbox(
                    f"Select from {len(matching_groups)} matching groups",
                    [""] + list(group_options.keys()),
                    key="admin_edit_group_select"
                )
                
                if selected_group_display:
                    selected_group_id = group_options[selected_group_display]
                    show_group_edit_form(selected_group_id)
            else:
                custom_info(f"No groups found matching '{search_term}'")
        except Exception as e:
            st.error(f"Error searching groups: {str(e)}")
    else:
        custom_info("Enter at least 3 characters in the search box to find groups to edit")

def show_group_edit_form(group_id: int):
    """Show the actual group editing form."""
    try:
        with get_db_session() as session:
            group_details = QuestionGroupService.get_group_details_with_verification(
                group_id=group_id, session=session
            )
        
        st.markdown(f"**Editing Group:** {group_details['display_title']}")
        
        # Basic group information editing
        new_display_title = st.text_input(
            "Group Display Title",
            value=group_details["display_title"],
            key="admin_edit_group_display_title"
        )
        edit_basic_col1, edit_basic_col2 = st.columns(2)
        with edit_basic_col1:
            new_is_reusable = st.checkbox(
                "Reusable across schemas",
                value=group_details["is_reusable"],
                key="admin_edit_group_reusable"
            )
        with edit_basic_col2:
            new_is_auto_submit = st.checkbox(
                "Auto Submit",
                value=group_details["is_auto_submit"],
                key="admin_edit_group_auto_submit"
            )
        
        new_description = st.text_area(
            "Description",
            value=group_details["description"] or "",
            key="admin_edit_question_group_description"
        )
        
        # Verification function editing
        current_verification = group_details.get("verification_function")
        st.markdown("**🔧 Verification Function:**")
        try:
            available_functions = QuestionGroupService.get_available_verification_functions()
            verification_options = ["None"] + available_functions
            current_index = 0
            if current_verification and current_verification in available_functions:
                current_index = verification_options.index(current_verification)
            
            new_verification_function = st.selectbox(
                "Select verification function",
                verification_options,
                index=current_index,
                key="admin_edit_group_verification",
                help="Optional function to validate answers"
            )
            
            new_verification_function = new_verification_function if new_verification_function != "None" else None
            
        except Exception as e:
            st.error(f"Error loading verification functions: {str(e)}")
            new_verification_function = current_verification
        
        # Question order management
        st.markdown("**📋 Question Order Management:**")
        with get_db_session() as session:
            current_order = QuestionGroupService.get_question_order(group_id=group_id, session=session)
        
        if current_order:
            order_key = f"edit_group_order_{group_id}"
            if order_key not in st.session_state:
                st.session_state[order_key] = current_order.copy()
            
            working_order = st.session_state[order_key]
            custom_info("💡 Use the ⬆️ and ⬇️ buttons to reorder questions. Changes will be applied when you click 'Update Group'.")
            
            # Load questions only when needed for the specific group
            with get_db_session() as session:
                questions_df = QuestionService.get_all_questions(session=session)
            
            for i, q_id in enumerate(working_order):
                question_row = questions_df[questions_df["ID"] == q_id]
                if not question_row.empty:
                    q_text = question_row.iloc[0]["Text"]
                    
                    order_col1, order_col2, order_col3 = st.columns([0.1, 0.8, 0.1])
                    
                    with order_col1:
                        if st.button("⬆️", key=f"up_{group_id}_{q_id}_{i}", 
                                    disabled=(i == 0), help="Move up"):
                            st.session_state[order_key][i], st.session_state[order_key][i-1] = \
                                st.session_state[order_key][i-1], st.session_state[order_key][i]
                            st.rerun(scope="fragment")
                    
                    with order_col2:
                        display_text = q_text[:80] + ('...' if len(q_text) > 80 else '')
                        st.write(f"**{i+1}.** {display_text}")
                        st.caption(f"ID: {q_id}")
                    
                    with order_col3:
                        if st.button("⬇️", key=f"down_{group_id}_{q_id}_{i}", 
                                    disabled=(i == len(working_order) - 1), help="Move down"):
                            st.session_state[order_key][i], st.session_state[order_key][i+1] = \
                                st.session_state[order_key][i+1], st.session_state[order_key][i]
                            st.rerun(scope="fragment")
            
            new_order = working_order
        else:
            new_order = current_order
            custom_info("No questions in this group.")
        
        # Update button with preview
        update_col, preview_col = st.columns(2)
        
        with update_col:
            if st.button("💾 Update Question Group", key="admin_update_group_btn", type="primary", use_container_width=True):
                try:
                    changes_made = []
                    
                    # Check what changed
                    if new_display_title != group_details["display_title"]:
                        changes_made.append("Display title")
                    
                    if new_description != (group_details["description"] or ""):
                        changes_made.append("Description")
                    
                    if new_is_reusable != group_details["is_reusable"]:
                        changes_made.append("Reusable setting")
                    
                    if new_is_auto_submit != group_details["is_auto_submit"]:
                        changes_made.append("Auto submit setting")
                    
                    if new_verification_function != current_verification:
                        changes_made.append("Verification function")
                    
                    if new_order != current_order:
                        changes_made.append("Question order")
                    
                    if changes_made:
                        with get_db_session() as session:
                            QuestionGroupService.edit_group(
                                group_id=group_id, new_display_title=new_display_title,
                                new_description=new_description, is_reusable=new_is_reusable,
                                verification_function=new_verification_function, is_auto_submit=new_is_auto_submit, session=session
                            )
                        
                        if new_order != current_order:
                            with get_db_session() as session:
                                QuestionGroupService.update_question_order(
                                    group_id=group_id, question_ids=new_order, session=session
                                )
                        
                        order_key = f"edit_group_order_{group_id}"
                        if order_key in st.session_state:
                            del st.session_state[order_key]
                        
                        custom_info(f"✅ Question group updated successfully! Changed: {', '.join(changes_made)}")
                        # st.rerun(scope="fragment")
                    else:
                        custom_info("No changes were made")
                        
                except Exception as e:
                    st.error(f"❌ Error updating group: {str(e)}")
        
        with preview_col:
            with st.expander("👁️ Preview Changes"):
                st.markdown("**Changes to be applied:**")
                
                if new_display_title != group_details["display_title"]:
                    st.markdown(f"**Display Title:** {group_details['display_title']} → {new_display_title}")
                
                if new_description != (group_details["description"] or ""):
                    st.markdown(f"**Description:** Updated")
                
                if new_is_reusable != group_details["is_reusable"]:
                    st.markdown(f"**Reusable:** {group_details['is_reusable']} → {new_is_reusable}")
                
                if new_is_auto_submit != group_details["is_auto_submit"]:
                    st.markdown(f"**Auto Submit:** {group_details['is_auto_submit']} → {new_is_auto_submit}")
                
                if new_verification_function != current_verification:
                    old_func = current_verification or "None"
                    new_func = new_verification_function or "None"
                    st.markdown(f"**Verification Function:** {old_func} → {new_func}")
                
                if new_order != current_order:
                    st.markdown(f"**Question Order:** Reordered")
                
                # Check if no changes
                no_changes = (
                    new_display_title == group_details["display_title"] and
                    new_description == (group_details["description"] or "") and
                    new_is_reusable == group_details["is_reusable"] and
                    new_is_auto_submit == group_details["is_auto_submit"] and
                    new_verification_function == current_verification and
                    new_order == current_order
                )
                
                if no_changes:
                    custom_info("No changes to preview")
                    
    except Exception as e:
        st.error(f"Error loading group details: {str(e)}")

def show_question_edit_interface():
    """Show question editing interface with search-first approach."""
    st.markdown("### ✏️ Edit Question")
    
    # Search-first approach
    search_term = st.text_input(
        "🔍 Search for question to edit", 
        placeholder="Enter question text to search...",
        key="admin_edit_question_search",
        help="Enter at least 3 characters to search"
    )
    
    if len(search_term) >= 3:
        try:
            # Search for questions matching the term
            with get_db_session() as session:
                matching_questions = QuestionService.search_questions_for_selection(
                    search_term=search_term, limit=20, session=session
                )
            
            if matching_questions:
                if len(matching_questions) == 20:
                    st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term.")
                else:
                    custom_info(f"✅ Found {len(matching_questions)} matching questions")
                
                # Show matching questions for selection
                question_options = {}
                for q in matching_questions:
                    archived_indicator = " 🗄️ (ARCHIVED)" if q['archived'] else ""
                    display_text = f"{q['text']}"
                    question_options[f"{display_text}{archived_indicator} (ID: {q['id']})"] = q['id']
                
                selected_question_display = st.selectbox(
                    f"Select from {len(matching_questions)} matching questions",
                    [""] + list(question_options.keys()),
                    key="admin_edit_question_select"
                )
                
                if selected_question_display:
                    selected_question_id = question_options[selected_question_display]
                    show_question_edit_form(selected_question_id)
            else:
                custom_info(f"No questions found matching '{search_term}'")
        except Exception as e:
            st.error(f"Error searching questions: {str(e)}")
    else:
        custom_info("Enter at least 3 characters in the search box to find questions to edit")


def show_question_edit_form(question_id: int):
    """Show the actual question editing form."""
    try:
        with get_db_session() as session:
            current_question = QuestionService.get_question_by_id(
                question_id=question_id, session=session
            )
        
        st.markdown(f"**Editing Question:** {current_question['text']}")
        
        st.text_input(
            "Question Text (immutable)",
            value=current_question["text"],
            key="admin_edit_question_text",
            disabled=True,
            help="Question text cannot be changed to preserve data integrity"
        )
        
        new_display_text = st.text_input(
            "Question to display to user",
            value=current_question["display_text"],
            key="admin_edit_question_display_text"
        )
        
        st.markdown(f"**Question Type:** `{current_question['type']}`")
        
        new_options = None
        new_default = None
        new_display_values = None
        new_option_weights = None
        
        if current_question["type"] == "single":
            st.markdown("**🎯 Options, Weights & Order Management:**")
            
            current_options = current_question["options"] or []
            current_display_values = current_question["display_values"] or current_options
            current_option_weights = current_question["option_weights"]
            if current_option_weights is None:
                current_option_weights = [1.0] * len(current_options)
            current_default = current_question["default_option"] or ""
            
            # Show current options
            st.markdown("**Current Options:**")
            for i, (opt, disp, weight) in enumerate(zip(current_options, current_display_values, current_option_weights)):
                default_indicator = " 🌟 (DEFAULT)" if opt == current_default else ""
                st.markdown(f"`{i+1}.` **Value:** `{opt}` | **Display:** `{disp}` | **Weight:** `{weight}`{default_indicator}")
            
            # Option editing (simplified version)
            st.markdown("**✏️ Edit Options:**")
            custom_info("📝 Note: You can only add new options, not remove existing ones (to preserve data integrity).")
            
            num_options = st.number_input(
                "Total number of options", 
                min_value=len(current_options), 
                max_value=10, 
                value=len(current_options),
                key="admin_edit_question_num_options"
            )
            
            new_options = []
            new_display_values = []
            new_option_weights = []
            
            for i in range(num_options):
                edit_opt_col1, edit_opt_col2, edit_opt_col3 = st.columns([2, 2, 1])
                
                with edit_opt_col1:
                    if i < len(current_options):
                        st.text_input(
                            f"Option {i+1} Value",
                            value=current_options[i],
                            disabled=True,
                            key=f"admin_edit_question_opt_val_{i}",
                            help="Cannot change existing option values"
                        )
                        new_options.append(current_options[i])
                    else:
                        new_opt = st.text_input(
                            f"Option {i+1} Value (NEW)",
                            key=f"admin_edit_question_opt_val_{i}",
                            placeholder="Enter new option value..."
                        )
                        if new_opt:
                            new_options.append(new_opt)
                
                with edit_opt_col2:
                    if i < len(current_display_values):
                        new_disp = st.text_input(
                            f"Option {i+1} Display",
                            value=current_display_values[i],
                            key=f"admin_edit_question_opt_disp_{i}"
                        )
                        new_display_values.append(new_disp if new_disp else current_display_values[i])
                    else:
                        new_disp = st.text_input(
                            f"Option {i+1} Display (NEW)",
                            value=new_options[i] if i < len(new_options) else "",
                            key=f"admin_edit_question_opt_disp_{i}",
                            placeholder="Display text for new option..."
                        )
                        if new_disp:
                            new_display_values.append(new_disp)
                        elif i < len(new_options):
                            new_display_values.append(new_options[i])
                
                with edit_opt_col3:
                    if i < len(current_option_weights):
                        new_weight = st.number_input(
                            f"Weight {i+1}",
                            min_value=0.0,
                            value=float(current_option_weights[i]),
                            step=0.1,
                            key=f"admin_edit_question_opt_weight_{i}",
                            help="Weight for scoring"
                        )
                        new_option_weights.append(new_weight)
                    else:
                        new_weight = st.number_input(
                            f"Weight {i+1} (NEW)",
                            min_value=0.0,
                            value=1.0,
                            step=0.1,
                            key=f"admin_edit_question_opt_weight_{i}",
                            help="Weight for scoring (default: 1.0)"
                        )
                        new_option_weights.append(new_weight)
            
            if new_options:
                new_default = st.selectbox(
                    "Default option",
                    [""] + new_options,
                    index=new_options.index(current_default) + 1 if current_default in new_options else 0,
                    key="admin_edit_question_default"
                )
                if new_default == "":
                    new_default = None
        
        elif current_question["type"] == "description":
            st.markdown("**📝 Description Question Settings:**")
            current_default = current_question.get("default_option", "")
            
            new_default = st.text_input(
                "Default response (optional)",
                value=current_default or "",
                key="admin_edit_question_default_desc",
                placeholder="Default text response...",
                help="Optional default response for description questions"
            )
            if not new_default.strip():
                new_default = None
        
        # Update button with preview
        update_col, preview_col = st.columns(2)
        
        with update_col:
            if st.button("💾 Update Question", key="admin_update_question_btn", type="primary", use_container_width=True):
                try:
                    changes_made = []
                    
                    # Check what changed
                    if new_display_text != current_question["display_text"]:
                        changes_made.append("Display text")
                    
                    if current_question["type"] == "single":
                        # Check for new options
                        if len(new_options) > len(current_options):
                            changes_made.append(f"Added {len(new_options) - len(current_options)} new options")
                        
                        # Check for display value changes
                        if new_display_values != current_display_values:
                            changes_made.append("Option display values")
                        
                        # Check for weight changes
                        if new_option_weights != current_option_weights:
                            changes_made.append("Option weights")
                        
                        # Check default change
                        if new_default != current_question["default_option"]:
                            changes_made.append("Default option")
                        
                        final_options = new_options
                        final_display_values = new_display_values
                        final_option_weights = new_option_weights
                    else:
                        # Description question
                        if new_default != current_question.get("default_option"):
                            changes_made.append("Default response")
                        
                        final_options = None
                        final_display_values = None
                        final_option_weights = None
                    
                    if changes_made:
                        with get_db_session() as session:
                            QuestionService.edit_question(
                                question_id=question_id, new_display_text=new_display_text,
                                new_opts=final_options, new_default=new_default,
                                new_display_values=final_display_values, new_option_weights=final_option_weights,
                                session=session
                            )
                        custom_info(f"✅ Question updated successfully! Changed: {', '.join(changes_made)}")
                        # st.rerun(scope="fragment")
                    else:
                        custom_info("No changes were made")
                        
                except Exception as e:
                    st.error(f"❌ Error updating question: {str(e)}")
        
        with preview_col:
            with st.expander("👁️ Preview Changes"):
                st.markdown("**Changes to be applied:**")
                
                if new_display_text != current_question["display_text"]:
                    st.markdown(f"**Display text:** Updated")
                
                if current_question["type"] == "single":
                    if len(new_options) > len(current_options):
                        st.markdown(f"**New options:** +{len(new_options) - len(current_options)}")
                    
                    if new_display_values != current_display_values:
                        st.markdown(f"**Option displays:** Updated")
                    
                    if new_option_weights != current_option_weights:
                        st.markdown(f"**Option weights:** Updated")
                    
                    if new_default != current_question["default_option"]:
                        old_default = current_question["default_option"] or "None"
                        new_default_display = new_default or "None"
                        st.markdown(f"**Default:** {old_default} → {new_default_display}")
                else:
                    if new_default != current_question.get("default_option"):
                        old_default = current_question.get("default_option") or "None"
                        new_default_display = new_default or "None"
                        st.markdown(f"**Default response:** {old_default} → {new_default_display}")
                
                # Check if no changes
                no_changes = (
                    new_display_text == current_question["display_text"] and
                    (current_question["type"] == "description" and new_default == current_question.get("default_option")) or
                    (current_question["type"] == "single" and 
                     len(new_options) == len(current_options) and
                     new_display_values == current_display_values and
                     new_option_weights == current_option_weights and
                     new_default == current_question["default_option"])
                )
                
                if no_changes:
                    custom_info("No changes to preview")
                    
    except Exception as e:
        st.error(f"Error loading question details: {str(e)}")


@st.fragment
def admin_export():
    st.subheader("📤 Ground Truth Export")
    
    # Get only counts for metrics - much faster
    try:
        with get_db_session() as session:
            export_counts = ProjectGroupService.get_export_counts(
                user_id=st.session_state.user["id"], 
                role="admin", 
                session=session
            )
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📁 Project Groups", export_counts["project_groups"])
        with col2:
            st.metric("📋 Total Projects", export_counts["total_projects"])
        with col3:
            st.metric("✅ Ready for Export", export_counts["ready_for_export"])
        with col4:
            selected_count = len(st.session_state.get("export_selected_projects", set()))
            st.metric("🎯 Selected", selected_count)
            
    except Exception as e:
        st.error(f"Error loading export summary: {str(e)}")
        export_counts = {"total_projects": 0}
    
    if export_counts["total_projects"] == 0:
        st.warning("No projects available for export.")
        return
    
    # Follow the same pattern as other tabs
    if st.button("📤 View & Select Projects for Export", use_container_width=True, type="primary"):
        clear_other_dialogs('export')
        st.rerun()
    
    if st.session_state.get('show_export_dialog', False):
        st.session_state['show_export_dialog'] = False
        show_export_dialog()


@st.fragment
@st.dialog("📤 Ground Truth Export - Project Selection", width="large")
def show_export_dialog():
    """Show export project selection interface in a dialog."""
    # Load the actual data
    try:
        grouped_projects = get_project_groups_with_projects(
            user_id=st.session_state.user["id"], 
            role="admin"
        )
        
        if not grouped_projects:
            st.warning("No projects available for export.")
            return
        
        show_export_project_interface(grouped_projects)
        
    except Exception as e:
        st.error(f"Error loading projects for export: {str(e)}")


def show_export_project_interface(grouped_projects):
    """Show the full export project selection interface."""
    
    # Initialize selection state
    if "export_selected_projects" not in st.session_state:
        st.session_state.export_selected_projects = set()
    
    # Search and filter
    st.markdown("### 🔍 Search & Filter")
    col1, col2, col3 = st.columns(3)
    with col1:
        search_term = st.text_input("🔍 Search projects", placeholder="Enter project name...")
    with col2:
        sort_by = st.selectbox("Sort by", ["Completion Rate", "Name"])
    with col3:
        sort_order = st.selectbox("Order", ["Ascending", "Descending"])
    
    st.markdown("---")
    
    # Track selections for batch updates (avoid individual reruns)
    current_selections = set(st.session_state.export_selected_projects)
    
    # Display project groups
    for group_index, (group_name, projects) in enumerate(grouped_projects.items()):
        if not projects:
            continue
        
        # Filter and enhance projects
        filtered_projects = [p for p in projects if not search_term or search_term.lower() in p["name"].lower()]
        if not filtered_projects:
            continue
        
        # Add completion info
        for project in filtered_projects:
            try:
                project["has_full_gt"] = check_project_has_full_ground_truth(project_id=project["id"])
                try:
                    with get_db_session() as session:
                        project_progress = ProjectService.progress(project_id=project["id"], session=session)
                    project["completion_rate"] = project_progress['completion_percentage']
                except:
                    project["completion_rate"] = 0.0
            except:
                project["has_full_gt"] = False
                project["completion_rate"] = 0.0
        
        # Sort projects
        if sort_by == "Completion Rate":
            filtered_projects.sort(key=lambda x: x["completion_rate"], reverse=(sort_order == "Descending"))
        else:
            filtered_projects.sort(key=lambda x: x["name"], reverse=(sort_order == "Descending"))
        
        # Pagination
        total_projects = len(filtered_projects)
        projects_per_page = 6
        total_pages = (total_projects - 1) // projects_per_page + 1 if total_projects > 0 else 1
        
        page_key = f"export_group_page_{group_name}"
        if search_term:
            current_page = 0
        else:
            if page_key not in st.session_state:
                st.session_state[page_key] = 0
            current_page = st.session_state[page_key]
            if current_page >= total_pages:
                current_page = 0
                st.session_state[page_key] = 0
        
        # Group header
        group_color = "#9553FE"
        selected_in_group = len([p for p in filtered_projects if p["id"] in current_selections])
        
        st.markdown(f"""
        <div style="{get_card_style(group_color)}position: relative;">
            <div style="position: absolute; top: -8px; left: 20px; background: {group_color}; color: white; padding: 4px 12px; border-radius: 10px; font-size: 0.8rem; font-weight: bold;">
                EXPORT GROUP
            </div>
            <div style="display: flex; justify-content: space-between; align-items: center; margin-top: 8px;">
                <div>
                    <h2 style="margin: 0; color: {group_color}; font-size: 1.8rem;">📁 {group_name}</h2>
                    <p style="margin: 8px 0 0 0; color: #34495e; font-size: 1.1rem; font-weight: 500;">
                        {selected_in_group}/{total_projects} selected {f"• Page {current_page + 1} of {total_pages}" if total_pages > 1 else ""}
                    </p>
                </div>
                <div style="text-align: right;">
                    <span style="background: {group_color}; color: white; padding: 10px 18px; border-radius: 20px; font-weight: bold;">{total_projects} Projects</span>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Four selection buttons - batch update to avoid multiple reruns
        btn_col1, btn_col2, btn_col3, btn_col4 = st.columns(4)
        start_idx = current_page * projects_per_page
        end_idx = min(start_idx + projects_per_page, total_projects)
        current_page_projects = filtered_projects[start_idx:end_idx]
        
        with btn_col1:
            if st.button("✅ Select Page", key=f"sel_pg_{group_name}_{current_page}", use_container_width=True):
                for p in current_page_projects:
                    current_selections.add(p["id"])
        with btn_col2:
            if st.button("❌ Deselect Page", key=f"desel_pg_{group_name}_{current_page}", use_container_width=True):
                for p in current_page_projects:
                    current_selections.discard(p["id"])
        with btn_col3:
            if st.button("✅ Select Group", key=f"sel_grp_{group_name}", use_container_width=True):
                for p in filtered_projects:
                    current_selections.add(p["id"])
        with btn_col4:
            if st.button("❌ Deselect Group", key=f"desel_grp_{group_name}", use_container_width=True):
                for p in filtered_projects:
                    current_selections.discard(p["id"])
        
        # Pagination controls with improved styling
        if total_pages > 1 and not search_term:
            page_col1, page_col2, page_col3 = st.columns([1, 1, 1])
            with page_col1:
                if st.button("◀ Previous", disabled=(current_page == 0), key=f"prev_exp_{group_name}", use_container_width=True):
                    st.session_state[page_key] = max(0, current_page - 1)
                    st.rerun(scope="fragment")
            with page_col2:
                # Custom page display - Page X / Y format
                page_options = [f"Page {i+1} / {total_pages}" for i in range(total_pages)]
                selected_page_display = st.selectbox(
                    "page_nav", 
                    page_options, 
                    index=current_page, 
                    key=f"pg_sel_exp_{group_name}",
                    label_visibility="collapsed"
                )
                new_page = page_options.index(selected_page_display)
                if new_page != current_page:
                    st.session_state[page_key] = new_page
                    st.rerun(scope="fragment")
            with page_col3:
                if st.button("Next ▶", disabled=(current_page == total_pages - 1), key=f"next_exp_{group_name}", use_container_width=True):
                    st.session_state[page_key] = min(total_pages - 1, current_page + 1)
                    st.rerun(scope="fragment")
        
        # Display projects with checkboxes (much simpler!)
        if current_page_projects:
            cols = st.columns(3)
            for i, project in enumerate(current_page_projects):
                with cols[i % 3]:
                    mode = "🎓 Training" if project["has_full_gt"] else "📝 Annotation"
                    completion_rate = project.get("completion_rate", 0.0)
                    status_text = "✅ Ready" if project["has_full_gt"] else "⚠️ Incomplete"
                    
                    # Simple card without style changes
                    st.markdown(f"""
                    <div style="border: 2px solid #9553FE; border-radius: 12px; padding: 18px; margin: 8px 0; 
                         background: linear-gradient(135deg, white, #f8f9fa); box-shadow: 0 4px 8px rgba(0,0,0,0.1); min-height: 200px;">
                        <h4 style="margin: 0 0 8px 0; color: black; font-size: 1.1rem;">{project['name']}</h4>
                        <p style="margin: 8px 0; color: #666; font-size: 0.9rem; min-height: 50px;">
                            {project.get("description", "") or 'No description'}
                        </p>
                        <div style="margin: 12px 0;">
                            <p style="margin: 4px 0;"><strong>Mode:</strong> {mode}</p>
                            <p style="margin: 4px 0;"><strong>Progress:</strong> {completion_rate:.1f}%</p>
                            <p style="margin: 4px 0; font-weight: bold;">{status_text}</p>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Simple checkbox - no rerun needed!
                    is_selected = st.checkbox(
                        "Select for export",
                        value=project["id"] in current_selections,
                        key=f"export_cb_{project['id']}",
                        label_visibility="visible"
                    )
                    
                    if is_selected:
                        current_selections.add(project["id"])
                    else:
                        current_selections.discard(project["id"])
        
        # Group separator
        if group_index < len(grouped_projects) - 1:
            st.markdown("""<div style="height: 2px; background: linear-gradient(90deg, transparent, #ddd, transparent); margin: 30px 0;"></div>""", unsafe_allow_html=True)
    
    # Update session state only once at the end (much faster!)
    if current_selections != st.session_state.export_selected_projects:
        st.session_state.export_selected_projects = current_selections
    
    # Move export configuration to bottom
    st.markdown("---")
    st.markdown("### ⚙️ Export Configuration")
    
    selected_count = len(current_selections)
    if selected_count == 0:
        st.warning("No projects selected for export.")
    else:
        # Show selected projects organized by groups (like assignment tab)
        selected_by_group = {}
        for group_name, projects in grouped_projects.items():
            selected_in_group = [p for p in projects if p["id"] in current_selections]
            if selected_in_group:
                selected_by_group[group_name] = [p["name"] for p in selected_in_group]
        
        if selected_by_group:
            custom_info(f"✅ <b>{selected_count}</b> projects selected")
            
            # Display selected projects organized by group
            for group_name, project_names in selected_by_group.items():
                if len(selected_by_group) > 1:  # Only show group headers if multiple groups
                    custom_info(f"📁 <b>{group_name}</b> ({len(project_names)} projects): {', '.join(project_names)}")
                else:
                    # If only one group, don't show group header
                    custom_info(f"📁 Selected {len(project_names)} projects from <b>{group_name}</b>: {', '.join(project_names)}")
        else:
            custom_info(f"✅ <b>{selected_count}</b> projects selected")
        
        config_col1, config_col2 = st.columns(2)
        with config_col1:
            export_format = st.selectbox("Format", ["json", "excel"], key="export_format")
        with config_col2:
            filename = st.text_input("Filename", value="ground_truth_export", key="export_filename")
        
        action_col1, action_col2, action_col3 = st.columns(3)
        
        with action_col1:
            if st.button("🔍 Validate", key="validate_btn", use_container_width=True):
                project_ids = list(current_selections)
                
                with st.spinner("Validating..."):
                    validation_result = validate_projects_for_export(project_ids)
                
                if validation_result["success"]:
                    custom_info("✅ Validation passed! Ready to export.")
                else:
                    st.error("❌ **Validation Failed**: Found conflicts in question group usage")
                    
                    # Display reusable errors
                    if validation_result["reusable_error"]:
                        st.markdown("**🔄 Reusable Question Group Conflicts:**")
                        reusable_groups = list(set([d["Group"] for d in validation_result["reusable_details"]]))
                        st.markdown(f"• **{len(reusable_groups)} reusable groups** with conflicting answers")
                        for group in reusable_groups[:3]:
                            questions_in_group = [d["Question"] for d in validation_result["reusable_details"] if d["Group"] == group]
                            st.markdown(f"  - **{group}**: {len(questions_in_group)} questions")
                        if len(reusable_groups) > 3:
                            st.markdown(f"  - ... and {len(reusable_groups) - 3} more groups")
                    
                    # Display non-reusable errors
                    if validation_result["non_reusable_error"]:
                        st.markdown("**🚫 Non-Reusable Question Group Violations:**")
                        non_reusable_groups = list(set([d["Group"] for d in validation_result["non_reusable_details"]]))
                        st.markdown(f"• **{len(non_reusable_groups)} non-reusable groups** used in multiple projects")
                        for group in non_reusable_groups[:3]:
                            st.markdown(f"  - **{group}**: Should only be used in one project")
                        if len(non_reusable_groups) > 3:
                            st.markdown(f"  - ... and {len(non_reusable_groups) - 3} more groups")
                    
                    # Overall summary
                    if validation_result["failing_videos"]:
                        failing_videos = validation_result["failing_videos"]
                        st.markdown(f"**📹 Total Affected Videos ({len(failing_videos)}):**")
                        video_display = ", ".join(failing_videos[:5])
                        if len(failing_videos) > 5:
                            video_display += f" ... and {len(failing_videos) - 5} more"
                        st.markdown(f"`{video_display}`")
                    
                    # Excel download
                    if validation_result["excel_report_data"] is not None:
                        try:
                            import io
                            buffer = io.BytesIO()
                            
                            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                validation_result["excel_report_data"].to_excel(writer, sheet_name='Validation Errors', index=False)
                                
                                worksheet = writer.sheets['Validation Errors']
                                
                                # Set column widths
                                worksheet.column_dimensions['A'].width = 25
                                worksheet.column_dimensions['B'].width = 35
                                worksheet.column_dimensions['C'].width = 100
                                worksheet.column_dimensions['D'].width = 70
                                worksheet.column_dimensions['E'].width = 18
                                worksheet.column_dimensions['F'].width = 20
                                worksheet.column_dimensions['G'].width = 15
                                
                                # Set row heights
                                for row in range(2, len(validation_result["excel_report_data"]) + 2):
                                    worksheet.row_dimensions[row].height = 120
                                
                                # Style headers
                                from openpyxl.styles import Font, PatternFill, Alignment
                                header_font = Font(bold=True, color="FFFFFF")
                                header_fill = PatternFill(start_color="9553FE", end_color="9553FE", fill_type="solid")
                                
                                for cell in worksheet[1]:
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                                
                                # Style data cells
                                wrap_alignment = Alignment(wrap_text=True, vertical="top")
                                for row in worksheet.iter_rows(min_row=2, max_row=len(validation_result["excel_report_data"]) + 1):
                                    for cell in row:
                                        cell.alignment = wrap_alignment
                            
                            buffer.seek(0)
                            
                            st.markdown("---")
                            st.download_button(
                                "📥 Download Comprehensive Error Report (Excel)",
                                buffer.getvalue(),
                                "validation_errors_comprehensive.xlsx", 
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                type="secondary"
                            )
                            
                            resolution_text = []
                            if validation_result["reusable_details"]:
                                resolution_text.append("**Reusable conflicts**: Align answers across projects")
                            if validation_result["non_reusable_details"]:
                                resolution_text.append("**Non-reusable violations**: Separate into different projects")
                            
                            custom_info(f"💡 **Next Steps**: {' and '.join(resolution_text)}. Download the comprehensive error report above for detailed guidance.")
                            
                        except Exception as excel_error:
                            st.warning(f"Could not generate Excel report: {str(excel_error)}")

        with action_col2:
            if st.button("📤 Export", key="export_btn", use_container_width=True, type="primary"):
                try:
                    project_ids = list(current_selections)
                    with st.spinner("Exporting..."):
                        with get_db_session() as session:
                            export_data = export_module.GroundTruthExportService.export_ground_truth_data(project_ids, session)
                        
                        file_ext = ".json" if export_format == "json" else ".xlsx"
                        final_filename = f"{filename}{file_ext}" if not filename.endswith(file_ext) else filename
                        
                        if export_format == "json":
                            import json
                            json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
                            st.download_button("📥 Download JSON", json_str, final_filename, "application/json", use_container_width=True)
                        else:
                            import io
                            buffer = io.BytesIO()
                            export_module.save_export_as_excel(export_data, buffer)
                            buffer.seek(0)
                            st.download_button("📥 Download Excel", buffer.getvalue(), final_filename, 
                                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    
                    custom_info(f"✅ Exported {len(export_data)} videos!")
                except ValueError as e:
                    st.error(f"❌ Export failed:\n\n{str(e)}")
        
        with action_col3:
            if st.button("🗑️ Clear All", key="clear_btn", use_container_width=True):
                st.session_state.export_selected_projects.clear()
                st.rerun(scope="fragment")

@st.fragment
def admin_users():
    st.subheader("👥 User Management")
    
    # Load only counts for metrics - much faster
    try:
        with get_db_session() as session:
            user_counts = AuthService.get_user_counts(session=session)
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("👥 Total Users", user_counts["total"])
        with col2:
            st.metric("✅ Active Users", user_counts["active"])
        with col3:
            st.metric("🗄️ Archived Users", user_counts["archived"])
        with col4:
            role_breakdown = f"👑{user_counts['admin']}/👤{user_counts['human']}/🤖{user_counts['model']}"
            st.metric("Admin/Human/Model", role_breakdown)
            
    except Exception as e:
        st.error(f"Error loading user summary: {str(e)}")
        user_counts = {"total": 0}
    
    # Option to view users table - now as a dialog
    if user_counts["total"] > 0:
        if st.button("👥 View All Users (Database Table)", use_container_width=True):
            clear_other_dialogs('users')
            st.rerun()
        
        if st.session_state.get('show_users_dialog', False):
            st.session_state['show_users_dialog'] = False
            show_users_table_dialog()
    
    # Management tabs
    user_management_tabs = st.tabs(["➕ Create User", "✏️ Edit User"])
    
    with user_management_tabs[0]:
        show_user_create_form()
    
    with user_management_tabs[1]:
        show_user_edit_interface()

def show_user_create_form():
    """Show user creation form."""
    st.markdown("### 🆕 Create New User")
    
    user_id = st.text_input("User ID", key="admin_user_id", placeholder="Enter unique user ID...")
    
    create_col1, create_col2 = st.columns(2)
    with create_col1:
        email = st.text_input("Email", key="admin_user_email", placeholder="user@example.com")
        user_type = st.selectbox("User Type", ["human", "model", "admin"], key="admin_user_type")
    with create_col2:
        password = st.text_input("Password", type="password", key="admin_user_password", placeholder="Enter password...")
        is_archived = st.checkbox("Create as archived", key="admin_user_archived", help="Create user in archived state")
    
    if st.button("🚀 Create User", key="admin_create_user_btn", type="primary", use_container_width=True):
        if user_id and password:
            # Email validation based on user type
            if user_type == "model" and email:
                st.error("❌ Model users cannot have emails")
                return
            elif user_type in ["human", "admin"] and not email:
                st.error("❌ Email is required for human and admin users")
                return
            
            try:
                with get_db_session() as session:
                    AuthService.create_user(
                        user_id=user_id, email=email, password_hash=password, 
                        user_type=user_type, is_archived=is_archived, session=session
                    )
                custom_info(f"✅ User '{user_id}' created successfully!")
                custom_info("💡 Refresh the page or navigate away and back to clear the form")
                # Removed manual session state clearing that was causing the error
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
        else:
            st.error("❌ User ID and Password are required")

def show_user_edit_interface():
    """Show user editing interface with search-first approach."""
    st.markdown("### ✏️ Edit User")
    
    # Search-first approach
    search_term = st.text_input(
        "🔍 Search for user to edit", 
        placeholder="Enter name or email to search...",
        key="admin_edit_user_search",
        help="Enter at least 3 characters to search"
    )
    
    filter_col1, filter_col2 = st.columns(2)
    with filter_col1:
        role_filter = st.selectbox("Filter by role", ["All", "admin", "human", "model"], key="admin_edit_user_role_filter")
    with filter_col2:
        show_archived = st.checkbox("Include archived users", value=True, key="admin_edit_user_show_archived")
    
    if len(search_term) >= 3:
        try:
            # Search for users matching the term
            with get_db_session() as session:
                matching_users = AuthService.search_users_for_selection(
                    search_term=search_term, user_role_filter=role_filter, limit=20, session=session
                )
            
            # Apply archived filter
            if not show_archived:
                matching_users = [u for u in matching_users if not u['archived']]
            
            if matching_users:
                if len(matching_users) == 20:
                    st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term.")
                else:
                    custom_info(f"✅ Found {len(matching_users)} matching users")
                
                # Show matching users for selection
                user_options = {}
                for u in matching_users:
                    archived_indicator = " 🗄️ (ARCHIVED)" if u['archived'] else ""
                    role_emoji = {"human": "👤", "admin": "👑", "model": "🤖"}.get(u['role'], "❓")
                    user_options[f"{role_emoji} {u['name']} ({u['email']}){archived_indicator} (ID: {u['id']})"] = u['id']
                
                selected_user_display = st.selectbox(
                    f"Select from {len(matching_users)} matching users",
                    [""] + list(user_options.keys()),
                    key="admin_edit_user_select"
                )
                
                if selected_user_display:
                    selected_user_id = user_options[selected_user_display]
                    show_user_edit_form(selected_user_id)
            else:
                custom_info(f"No users found matching '{search_term}' with the selected filters")
        except Exception as e:
            st.error(f"Error searching users: {str(e)}")
    else:
        custom_info("Enter at least 3 characters in the search box to find users to edit")

def show_user_edit_form(user_id: int):
    """Show the actual user editing form."""
    try:
        with get_db_session() as session:
            current_user_info = AuthService.get_user_info_by_id(user_id=user_id, session=session)
        
        st.markdown(f"**Editing User:** {current_user_info['user_id_str']}")
        
        # Basic user information editing
        col1, col2 = st.columns(2)
        
        with col1:
            new_user_id = st.text_input(
                "User ID",
                value=current_user_info['user_id_str'],
                key="admin_edit_user_id"
            )
            new_email = st.text_input(
                "Email",
                value=current_user_info['email'] or "",
                key="admin_edit_user_email",
                help="Model users cannot have emails"
            )
        
        with col2:
            new_role = st.selectbox(
                "User Type",
                ["human", "model", "admin"],
                index=["human", "model", "admin"].index(current_user_info['user_type']),
                key="admin_edit_user_role"
            )
            new_password = st.text_input(
                "New Password (leave empty to keep current)",
                type="password",
                key="admin_edit_user_password"
            )
        
        # Archive status
        new_archived = st.checkbox(
            "Archived",
            value=current_user_info['is_archived'],
            key="admin_edit_user_archived",
            help="Archive user to prevent login and assignments"
        )
        
        # Update button with preview
        update_col, preview_col = st.columns(2)
        
        with update_col:
            if st.button("💾 Update User", key="admin_update_user_btn", type="primary", use_container_width=True):
                try:
                    changes_made = []
                    
                    # Validate email based on role
                    if new_role == "model" and new_email:
                        st.error("❌ Model users cannot have emails")
                        return
                    elif new_role in ["human", "admin"] and not new_email:
                        st.error("❌ Email is required for human and admin users")
                        return
                    
                    with get_db_session() as session:
                        # Check what changed
                        if new_user_id != current_user_info['user_id_str']:
                            AuthService.update_user_id(
                                user_id=user_id, new_user_id=new_user_id, session=session
                            )
                            changes_made.append("User ID")
                        
                        if new_email != (current_user_info['email'] or ""):
                            AuthService.update_user_email(
                                user_id=user_id, new_email=new_email, session=session
                            )
                            changes_made.append("Email")
                        
                        if new_password:
                            AuthService.update_user_password(
                                user_id=user_id, new_password=new_password, session=session
                            )
                            changes_made.append("Password")
                        
                        if new_role != current_user_info['user_type']:
                            AuthService.update_user_role(
                                user_id=user_id, new_role=new_role, session=session
                            )
                            changes_made.append("Role")
                        
                        if new_archived != current_user_info['is_archived']:
                            AuthService.toggle_user_archived(user_id=user_id, session=session)
                            changes_made.append("Archive status")
                    
                    if changes_made:
                        custom_info(f"✅ User '{new_user_id}' updated successfully! Changed: {', '.join(changes_made)}")
                    else:
                        custom_info("No changes were made")
                        
                except Exception as e:
                    st.error(f"❌ Error updating user: {str(e)}")
        
        with preview_col:
            with st.expander("👁️ Preview Changes"):
                st.markdown("**Changes to be applied:**")
                
                if new_user_id != current_user_info['user_id_str']:
                    st.markdown(f"**User ID:** {current_user_info['user_id_str']} → {new_user_id}")
                
                if new_email != (current_user_info['email'] or ""):
                    old_email = current_user_info['email'] or "None"
                    st.markdown(f"**Email:** {old_email} → {new_email}")
                
                if new_password:
                    st.markdown(f"**Password:** Will be updated")
                
                if new_role != current_user_info['user_type']:
                    st.markdown(f"**Role:** {current_user_info['user_type']} → {new_role}")
                
                if new_archived != current_user_info['is_archived']:
                    status_change = "Archived" if new_archived else "Unarchived"
                    st.markdown(f"**Status:** {status_change}")
                
                # Check if no changes
                no_changes = (
                    new_user_id == current_user_info['user_id_str'] and
                    new_email == (current_user_info['email'] or "") and
                    not new_password and
                    new_role == current_user_info['user_type'] and
                    new_archived == current_user_info['is_archived']
                )
                
                if no_changes:
                    custom_info("No changes to preview")
                    
    except Exception as e:
        st.error(f"Error loading user details: {str(e)}")



@st.fragment 
def admin_project_groups():
    st.subheader("📊 Project Group Management")
    
    # Load only counts for metrics - much faster
    try:
        # Get optimized counts
        with get_db_session() as session:
            group_counts = ProjectGroupService.get_project_group_counts(session=session)
            project_counts = ProjectService.get_project_counts(session=session)
        
        # Calculate ungrouped projects
        ungrouped_projects = project_counts["total"] - group_counts["total_projects_in_groups"]
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📊 Total Groups", group_counts["total_groups"])
        with col2:
            st.metric("📁 Projects in Groups", group_counts["total_projects_in_groups"])
        with col3:
            st.metric("📈 Avg Projects/Group", group_counts["avg_projects_per_group"])
        with col4:
            st.metric("📭 Ungrouped Projects", ungrouped_projects)
            
    except Exception as e:
        st.error(f"Error loading project group summary: {str(e)}")
        group_counts = {"total_groups": 0}
    
    # Option to view groups table - now as a dialog
    if group_counts["total_groups"] > 0:
        if st.button("📊 View All Project Groups (Database Table)", use_container_width=True):
            clear_other_dialogs('project_groups')
            st.rerun()
        
        if st.session_state.get('show_project_groups_dialog', False):
            st.session_state['show_project_groups_dialog'] = False
            show_project_groups_table_dialog()
    
    # Management tabs
    group_management_tabs = st.tabs(["➕ Create Group", "✏️ Edit Group"])
    
    with group_management_tabs[0]:
        show_project_group_create_form()
    
    with group_management_tabs[1]:
        show_project_group_edit_interface()

def show_project_group_create_form():
    """Show project group creation form with search-optimized project selection."""
    st.markdown("### 🆕 Create New Project Group")
    
    group_name = st.text_input("Group Name", key="admin_pgroup_name", placeholder="Enter group name...")
    group_description = st.text_area("Description", key="admin_pgroup_description", placeholder="Describe the purpose of this group...")
    
    # Project selection with search-first approach
    st.markdown("**📁 Select Projects (Optional):**")
    project_search = st.text_input("🔍 Search projects to add", placeholder="Enter project name (min 3 chars)...", key="admin_pgroup_project_search", help="Enter at least 3 characters to search")
    
    # Initialize selected projects in session state - NOW STORE OBJECTS WITH ID AND NAME
    if "create_pgroup_selected_projects" not in st.session_state:
        st.session_state.create_pgroup_selected_projects = []  # List of {"id": int, "name": str}
    
    # Show search results
    if len(project_search) >= 3:
        try:
            with get_db_session() as session:
                matching_projects = ProjectService.search_projects_for_selection(search_term=project_search, limit=20, session=session)
            
            if matching_projects:
                if len(matching_projects) == 20:
                    st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term to find your projects faster.")
                else:
                    custom_info(f"✅ Found {len(matching_projects)} matching projects")
                
                for project in matching_projects:
                    project_col1, project_col2 = st.columns([1, 3])
                    
                    with project_col1:
                        # Check if already selected using ID
                        selected_ids = [p["id"] for p in st.session_state.create_pgroup_selected_projects]
                        is_selected = project['id'] in selected_ids
                        
                        if st.checkbox(f"Select", value=is_selected, key=f"project_select_{project['id']}"):
                            if project['id'] not in selected_ids:
                                # Store full object with ID and name
                                st.session_state.create_pgroup_selected_projects.append({
                                    "id": project['id'],
                                    "name": project['name']
                                })
                        else:
                            # Remove from selection
                            st.session_state.create_pgroup_selected_projects = [
                                p for p in st.session_state.create_pgroup_selected_projects 
                                if p["id"] != project['id']
                            ]
                    
                    with project_col2:
                        archived_indicator = " 🗄️ (ARCHIVED)" if project.get('archived', False) else ""
                        st.markdown(f"**{project['name']}**{archived_indicator} (ID: {project['id']})")
            else:
                custom_info(f"No projects found matching '{project_search}'. Try a different search term.")
        except Exception as e:
            st.error(f"Error searching projects: {str(e)}")
    else:
        custom_info("Enter at least 3 characters in the search box to find projects, or leave empty to create group without projects")
    
    # Show selected projects - NOW WITH PROPER NAMES
    if st.session_state.create_pgroup_selected_projects:
        st.markdown(f"**Selected {len(st.session_state.create_pgroup_selected_projects)} projects:**")
        
        selected_col1, selected_col2 = st.columns([1, 3])
        
        with selected_col1:
            if st.button("Clear All", key="clear_selected_projects_pgroup"):
                st.session_state.create_pgroup_selected_projects = []
                st.rerun()
        
        with selected_col2:
            for project_obj in st.session_state.create_pgroup_selected_projects[:10]:
                # NOW SHOW ACTUAL NAMES!
                st.text(f"• {project_obj['name']}")
            if len(st.session_state.create_pgroup_selected_projects) > 10:
                st.caption(f"... and {len(st.session_state.create_pgroup_selected_projects) - 10} more")
    
    if st.button("🚀 Create Project Group", key="admin_create_pgroup_btn", type="primary", use_container_width=True):
        if group_name:
            try:
                # Extract just IDs for the service call
                selected_project_ids = [p["id"] for p in st.session_state.create_pgroup_selected_projects]
                
                with get_db_session() as session:
                    ProjectGroupService.create_project_group(
                        name=group_name, description=group_description, 
                        project_ids=selected_project_ids, session=session
                    )
                
                # Clear selection after successful creation
                st.session_state.create_pgroup_selected_projects = []
                custom_info("✅ Project group created successfully!")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
        else:
            st.error("❌ Group name is required")

def show_project_group_edit_interface():
    """Show project group editing interface with search-first approach."""
    st.markdown("### ✏️ Edit Project Group")
    
    # Search-first approach
    search_term = st.text_input(
        "🔍 Search for group to edit", 
        placeholder="Enter group name to search...",
        key="admin_edit_pgroup_search",
        help="Enter at least 3 characters to search"
    )
    
    if len(search_term) >= 3:
        try:
            # Search for groups matching the term
            with get_db_session() as session:
                groups = ProjectGroupService.list_project_groups(session=session)
            matching_groups = [g for g in groups if search_term.lower() in g['name'].lower()]
            
            if matching_groups:
                if len(matching_groups) > 20:
                    st.warning(f"⚠️ Found {len(matching_groups)} matches. **Too many results!** Try a more specific search term.")
                    matching_groups = matching_groups[:20]
                else:
                    custom_info(f"✅ Found {len(matching_groups)} matching groups")
                
                # Show matching groups for selection
                group_options = {f"{g['name']} (ID: {g['id']})": g['id'] for g in matching_groups}
                
                selected_group_display = st.selectbox(
                    f"Select from {len(matching_groups)} matching groups",
                    [""] + list(group_options.keys()),
                    key="admin_edit_pgroup_select"
                )
                
                if selected_group_display:
                    selected_group_id = group_options[selected_group_display]
                    show_project_group_edit_form(selected_group_id)
            else:
                custom_info(f"No groups found matching '{search_term}'")
        except Exception as e:
            st.error(f"Error searching groups: {str(e)}")
    else:
        custom_info("Enter at least 3 characters in the search box to find groups to edit")

def show_project_group_edit_form(group_id: int):
    """Show the actual project group editing form."""
    try:
        with get_db_session() as session:
            group_info = ProjectGroupService.get_project_group_by_id(
                group_id=group_id, session=session
            )
        current_group = group_info["group"]
        current_projects = group_info["projects"]
        
        st.markdown(f"**Editing Group:** {current_group['name']}")
        
        # Basic group information editing
        new_name = st.text_input(
            "Group Name", 
            value=current_group["name"],
            key="admin_edit_group_name"
        )
        new_description = st.text_area(
            "Description", 
            value=current_group["description"] or "",
            key="admin_edit_pgroup_description"
        )
        
        # Project management with search-first approach
        st.markdown("**📁 Project Management:**")
        
        if current_projects:
            st.markdown("**Current Projects:**")
            current_project_ids = [p["id"] for p in current_projects]
            for project in current_projects:
                archived_indicator = " 🗄️" if project.get("is_archived", False) else " ✅"
                st.write(f"- {project['name']}{archived_indicator} (ID: {project['id']})")
        else:
            custom_info("No projects currently in this group")
            current_project_ids = []
        
        # Initialize session state for project selections
        if f"edit_pgroup_add_projects_{group_id}" not in st.session_state:
            st.session_state[f"edit_pgroup_add_projects_{group_id}"] = []
        if f"edit_pgroup_remove_projects_{group_id}" not in st.session_state:
            st.session_state[f"edit_pgroup_remove_projects_{group_id}"] = []
        
        # Add Projects Section with Search
        st.markdown("**➕ Add Projects to Group:**")
        add_search = st.text_input(
            "🔍 Search projects to add", 
            placeholder="Enter project name (min 3 chars)...", 
            key=f"admin_edit_pgroup_add_search_{group_id}",
            help="Search for projects not currently in this group"
        )
        
        if len(add_search) >= 3:
            try:
                # Get projects not currently in the group
                with get_db_session() as session:
                    all_projects_df = ProjectService.get_all_projects_including_archived(session=session)
                available_to_add = all_projects_df[~all_projects_df["ID"].isin(current_project_ids)]
                
                if not available_to_add.empty:
                    # Filter by search term
                    filtered_add_projects = available_to_add[
                        available_to_add["Name"].str.contains(add_search, case=False, na=False)
                    ]
                    
                    if not filtered_add_projects.empty:
                        if len(filtered_add_projects) >= 20:
                            st.warning(f"⚠️ Showing first 20 results. **Too many matches!** Try a more specific search term.")
                            filtered_add_projects = filtered_add_projects.head(20)
                        else:
                            custom_info(f"✅ Found {len(filtered_add_projects)} matching projects to add")
                        
                        for _, project_row in filtered_add_projects.iterrows():
                            project_col1, project_col2 = st.columns([1, 3])
                            
                            with project_col1:
                                selected_add_ids = [p["id"] for p in st.session_state[f"edit_pgroup_add_projects_{group_id}"]]
                                is_selected = project_row["ID"] in selected_add_ids
                                
                                if st.checkbox(f"Add", value=is_selected, key=f"add_project_select_{group_id}_{project_row['ID']}"):
                                    if project_row["ID"] not in selected_add_ids:
                                        st.session_state[f"edit_pgroup_add_projects_{group_id}"].append({
                                            "id": project_row["ID"],
                                            "name": project_row["Name"]
                                        })
                                else:
                                    st.session_state[f"edit_pgroup_add_projects_{group_id}"] = [
                                        p for p in st.session_state[f"edit_pgroup_add_projects_{group_id}"] 
                                        if p["id"] != project_row["ID"]
                                    ]
                            
                            with project_col2:
                                archived_indicator = " 🗄️ (ARCHIVED)" if project_row["Archived"] else ""
                                st.markdown(f"**{project_row['Name']}**{archived_indicator} (ID: {project_row['ID']})")
                    else:
                        custom_info(f"No available projects found matching '{add_search}'")
                else:
                    custom_info("All projects are already in this group")
            except Exception as e:
                st.error(f"Error searching projects to add: {str(e)}")
        else:
            custom_info("Enter at least 3 characters to search for projects to add")
        
        # Show selected projects to add
        if st.session_state[f"edit_pgroup_add_projects_{group_id}"]:
            st.markdown(f"**Projects to Add ({len(st.session_state[f'edit_pgroup_add_projects_{group_id}'])}):**")
            
            add_col1, add_col2 = st.columns([1, 3])
            
            with add_col1:
                if st.button("Clear Add Selection", key=f"clear_add_projects_{group_id}"):
                    st.session_state[f"edit_pgroup_add_projects_{group_id}"] = []
                    st.rerun()
            
            with add_col2:
                for project_obj in st.session_state[f"edit_pgroup_add_projects_{group_id}"][:10]:
                    st.text(f"• {project_obj['name']}")
                if len(st.session_state[f"edit_pgroup_add_projects_{group_id}"]) > 10:
                    st.caption(f"... and {len(st.session_state[f'edit_pgroup_add_projects_{group_id}']) - 10} more")
        
        # Remove Projects Section with Search (only if there are current projects)
        if current_projects:
            st.markdown("**➖ Remove Projects from Group:**")
            remove_search = st.text_input(
                "🔍 Search projects to remove", 
                placeholder="Enter project name (min 3 chars)...", 
                key=f"admin_edit_pgroup_remove_search_{group_id}",
                help="Search within current group projects"
            )
            
            if len(remove_search) >= 3:
                # Filter current projects by search term
                filtered_remove_projects = [
                    p for p in current_projects 
                    if remove_search.lower() in p["name"].lower()
                ]
                
                if filtered_remove_projects:
                    custom_info(f"✅ Found {len(filtered_remove_projects)} matching projects to remove")
                    
                    for project in filtered_remove_projects:
                        project_col1, project_col2 = st.columns([1, 3])
                        
                        with project_col1:
                            selected_remove_ids = [p["id"] for p in st.session_state[f"edit_pgroup_remove_projects_{group_id}"]]
                            is_selected = project["id"] in selected_remove_ids
                            
                            if st.checkbox(f"Remove", value=is_selected, key=f"remove_project_select_{group_id}_{project['id']}"):
                                if project["id"] not in selected_remove_ids:
                                    st.session_state[f"edit_pgroup_remove_projects_{group_id}"].append({
                                        "id": project["id"],
                                        "name": project["name"]
                                    })
                            else:
                                st.session_state[f"edit_pgroup_remove_projects_{group_id}"] = [
                                    p for p in st.session_state[f"edit_pgroup_remove_projects_{group_id}"] 
                                    if p["id"] != project["id"]
                                ]
                        
                        with project_col2:
                            archived_indicator = " 🗄️" if project.get("is_archived", False) else " ✅"
                            st.markdown(f"**{project['name']}**{archived_indicator} (ID: {project['id']})")
                else:
                    custom_info(f"No current projects found matching '{remove_search}'")
            else:
                custom_info("Enter at least 3 characters to search for projects to remove")
            
            # Show selected projects to remove
            if st.session_state[f"edit_pgroup_remove_projects_{group_id}"]:
                st.markdown(f"**Projects to Remove ({len(st.session_state[f'edit_pgroup_remove_projects_{group_id}'])}):**")
                
                remove_col1, remove_col2 = st.columns([1, 3])
                
                with remove_col1:
                    if st.button("Clear Remove Selection", key=f"clear_remove_projects_{group_id}"):
                        st.session_state[f"edit_pgroup_remove_projects_{group_id}"] = []
                        st.rerun()
                
                with remove_col2:
                    for project_obj in st.session_state[f"edit_pgroup_remove_projects_{group_id}"][:10]:
                        st.text(f"• {project_obj['name']}")
                    if len(st.session_state[f"edit_pgroup_remove_projects_{group_id}"]) > 10:
                        st.caption(f"... and {len(st.session_state[f'edit_pgroup_remove_projects_{group_id}']) - 10} more")
        
        # Extract IDs for the service call
        add_projects = [p["id"] for p in st.session_state[f"edit_pgroup_add_projects_{group_id}"]]
        remove_projects = [p["id"] for p in st.session_state[f"edit_pgroup_remove_projects_{group_id}"]]
        
        # Update button with preview
        update_col, preview_col = st.columns(2)
        
        with update_col:
            if st.button("💾 Update Project Group", key="admin_update_pgroup_btn", type="primary", use_container_width=True):
                try:
                    changes_made = []
                    
                    # Check what changed
                    if new_name != current_group["name"]:
                        changes_made.append("Name")
                    
                    if new_description != (current_group["description"] or ""):
                        changes_made.append("Description")
                    
                    if add_projects:
                        changes_made.append(f"Added {len(add_projects)} projects")
                    
                    if remove_projects:
                        changes_made.append(f"Removed {len(remove_projects)} projects")
                    
                    if changes_made:
                        with get_db_session() as session:
                            ProjectGroupService.edit_project_group(
                                group_id=group_id,
                                name=new_name if new_name != current_group["name"] else None,
                                description=new_description if new_description != (current_group["description"] or "") else None,
                                add_project_ids=add_projects if add_projects else None,
                                remove_project_ids=remove_projects if remove_projects else None,
                                session=session
                            )
                        
                        # Clear selections after successful update
                        st.session_state[f"edit_pgroup_add_projects_{group_id}"] = []
                        st.session_state[f"edit_pgroup_remove_projects_{group_id}"] = []
                        
                        custom_info(f"✅ Project group updated successfully! Changed: {', '.join(changes_made)}")
                    else:
                        custom_info("No changes were made")
                        
                except Exception as e:
                    st.error(f"❌ Error updating project group: {str(e)}")
        
        with preview_col:
            with st.expander("👁️ Preview Changes"):
                st.markdown("**Changes to be applied:**")
                
                if new_name != current_group["name"]:
                    st.markdown(f"**Name:** {current_group['name']} → {new_name}")
                
                if new_description != (current_group["description"] or ""):
                    st.markdown(f"**Description:** Updated")
                
                if add_projects:
                    st.markdown(f"**Add Projects:** {len(add_projects)} projects")
                    for project_obj in st.session_state[f"edit_pgroup_add_projects_{group_id}"][:5]:
                        st.markdown(f"  • {project_obj['name']}")
                    if len(st.session_state[f"edit_pgroup_add_projects_{group_id}"]) > 5:
                        st.markdown(f"  • ... and {len(st.session_state[f'edit_pgroup_add_projects_{group_id}']) - 5} more")
                
                if remove_projects:
                    st.markdown(f"**Remove Projects:** {len(remove_projects)} projects")
                    for project_obj in st.session_state[f"edit_pgroup_remove_projects_{group_id}"][:5]:
                        st.markdown(f"  • {project_obj['name']}")
                    if len(st.session_state[f"edit_pgroup_remove_projects_{group_id}"]) > 5:
                        st.markdown(f"  • ... and {len(st.session_state[f'edit_pgroup_remove_projects_{group_id}']) - 5} more")
                
                # Check if no changes
                no_changes = (
                    new_name == current_group["name"] and
                    new_description == (current_group["description"] or "") and
                    not add_projects and
                    not remove_projects
                )
                
                if no_changes:
                    custom_info("No changes to preview")
                    
    except Exception as e:
        st.error(f"Error loading project group details: {str(e)}")

###############################################################################
# Admin Helper Functions
###############################################################################

def validate_projects_for_export(project_ids: List[int]):
    """
    Validate projects for export and return structured results.
    
    Returns:
        Dict with keys: success, reusable_error, non_reusable_error, failing_videos, 
                       reusable_details, non_reusable_details, excel_report_data
    """
    reusable_error = None
    non_reusable_error = None
    
    with get_db_session() as session:
        # Check reusable question groups
        try:
            export_module.GroundTruthExportService._validate_reusable_question_groups(project_ids, session)
        except ValueError as e:
            reusable_error = str(e)
        
        # Check non-reusable question groups  
        try:
            export_module.GroundTruthExportService._validate_non_reusable_question_groups(project_ids, session)
        except ValueError as e:
            non_reusable_error = str(e)
    
    # If no errors, return success
    if not reusable_error and not non_reusable_error:
        return {"success": True}
    
    # Parse errors
    all_failing_videos = set()
    reusable_details = []
    non_reusable_details = []
    
    # Parse reusable errors
    if reusable_error:
        if "Failing videos:" in reusable_error:
            videos_section = reusable_error.split("Failing videos:")[1].split("Details:")[0].strip()
            failing_videos = [v.strip() for v in videos_section.replace("...", "").split(",") if v.strip()]
            all_failing_videos.update(failing_videos)
        
        if "Details:" in reusable_error:
            details_section = reusable_error.split("Details:")[1].strip()
            current_group = None
            
            for line in details_section.split("\n"):
                line = line.strip()
                if "Reusable question group" in line and "failing videos" in line:
                    current_group = line.split("'")[1] if "'" in line else "Unknown Group"
                elif "Question '" in line and current_group:
                    question = line.split("Question '")[1].split("':")[0] if "Question '" in line else "Unknown Question"
                    reusable_details.append({
                        "Type": "Reusable Group Conflict",
                        "Group": current_group,
                        "Question": question,
                        "Issue": "Inconsistent answers across projects"
                    })
    
    # Parse non-reusable errors
    if non_reusable_error:
        if "Total videos affected:" in non_reusable_error:
            lines = non_reusable_error.split("\n")
            for line in lines:
                if "Overlapping videos" in line and "):" in line:
                    videos_part = line.split("):")[1] if "):" in line else ""
                    failing_videos = [v.strip() for v in videos_part.replace("...", "").split(",") if v.strip()]
                    all_failing_videos.update(failing_videos)
        
        if "Violations:" in non_reusable_error:
            violations_section = non_reusable_error.split("Violations:")[1].strip()
            current_group = None
            
            for line in violations_section.split("\n"):
                line = line.strip()
                if "Non-reusable question group" in line:
                    current_group = line.split("'")[1] if "'" in line else "Unknown Group"
                elif "Projects:" in line and current_group:
                    projects_part = line.split("Projects:")[1].strip()
                    current_projects = [p.strip() for p in projects_part.split(",")]
                    non_reusable_details.append({
                        "Type": "Non-Reusable Group Violation", 
                        "Group": current_group,
                        "Question": "N/A",
                        "Issue": f"Group used in multiple projects: {', '.join(current_projects)}"
                    })
    
    # Create Excel report data
    excel_report_data = None
    all_failing_videos = list(all_failing_videos)
    
    if (reusable_details or non_reusable_details) and all_failing_videos:
        try:
            import pandas as pd
            
            # Group errors by video
            video_errors = {}
            for video in all_failing_videos:
                video_errors[video] = {"reusable": [], "non_reusable": []}
            
            # Add reusable errors
            for detail in reusable_details:
                error_msg = f"Group '{detail['Group']}' → Question '{detail['Question']}': {detail['Issue']}"
                for video in all_failing_videos:
                    video_errors[video]["reusable"].append(error_msg)
            
            # Add non-reusable errors
            for detail in non_reusable_details:
                error_msg = f"Group '{detail['Group']}': {detail['Issue']}"
                for video in all_failing_videos:
                    video_errors[video]["non_reusable"].append(error_msg)
            
            # Create report rows
            report_data = []
            for video, errors in video_errors.items():
                reusable_errors = errors["reusable"]
                non_reusable_errors = errors["non_reusable"]
                
                # Combine all errors
                all_errors = []
                if reusable_errors:
                    all_errors.append("REUSABLE GROUP CONFLICTS:")
                    all_errors.extend([f"• {err}" for err in reusable_errors])
                if non_reusable_errors:
                    if all_errors:
                        all_errors.append("\nNON-REUSABLE GROUP VIOLATIONS:")
                    else:
                        all_errors.append("NON-REUSABLE GROUP VIOLATIONS:")
                    all_errors.extend([f"• {err}" for err in non_reusable_errors])
                
                combined_errors = "\n".join(all_errors)
                
                # Determine issue types
                issue_types = []
                if reusable_errors:
                    issue_types.append("Reusable Group Conflicts")
                if non_reusable_errors:
                    issue_types.append("Non-Reusable Group Violations")
                
                # Create resolution guidance
                resolution_steps = []
                if reusable_errors:
                    resolution_steps.append("1. Review conflicting answers in reusable question groups across projects")
                    resolution_steps.append("2. Align answers to be consistent for the same reusable groups")
                if non_reusable_errors:
                    resolution_steps.append("3. Move non-reusable question groups to separate projects")
                    resolution_steps.append("4. Or convert to reusable groups if they should be shared")
                
                report_data.append({
                    "Video ID": video,
                    "Issue Types": " + ".join(issue_types),
                    "Detailed Errors": combined_errors,
                    "Resolution Steps": "\n".join(resolution_steps),
                    "Reusable Conflicts": len(reusable_errors),
                    "Non-Reusable Violations": len(non_reusable_errors),
                    "Total Issues": len(reusable_errors) + len(non_reusable_errors)
                })
            
            excel_report_data = pd.DataFrame(report_data)
            
        except Exception as e:
            print(f"Error creating Excel report data: {e}")
    
    return {
        "success": False,
        "reusable_error": reusable_error,
        "non_reusable_error": non_reusable_error,
        "failing_videos": all_failing_videos,
        "reusable_details": reusable_details,
        "non_reusable_details": non_reusable_details,
        "excel_report_data": excel_report_data
    }

      

def clear_other_dialogs(current_dialog):
    """Clear all dialog flags except the current one."""
    dialog_flags = [
        'show_videos_dialog',
        'show_projects_dialog', 
        'show_schemas_dialog',
        'show_questions_dialog',
        'show_question_groups_dialog',
        'show_project_groups_dialog',
        'show_users_dialog',
        'show_assignments_dialog',
        'show_export_dialog',
    ]
    
    for flag in dialog_flags:
        st.session_state[flag] = False
    
    # Set the current dialog flag
    st.session_state[f'show_{current_dialog}_dialog'] = True

@st.fragment
@st.dialog("📋 Videos Database Table", width="large")
def show_videos_table_dialog():
    """Show videos table in a dialog with search and pagination."""
    # Search and filter controls
    search_col1, search_col2, search_col3 = st.columns([2, 1, 1])
    with search_col1:
        video_search = st.text_input("🔍 Search videos", placeholder="Video UID, URL...", key="videos_dialog_video_search")
    with search_col2:
        show_archived = st.checkbox("Show archived", key="videos_dialog_show_archived")
    with search_col3:
        show_only_unassigned = st.checkbox("Only unassigned", key="videos_dialog_show_unassigned")
    
    # Pagination controls
    page_size = st.selectbox("Videos per page", [25, 50, 100], index=1, key="videos_dialog_page_size")
    
    # Initialize page in session state
    if "videos_dialog_page" not in st.session_state:
        st.session_state.videos_dialog_page = 0
    
    try:
        # Search videos with pagination
        with get_db_session() as session:
            result = VideoService.search_videos(
                search_term=video_search,
                show_archived=show_archived,
                show_only_unassigned=show_only_unassigned,
                page=st.session_state.videos_dialog_page,
                page_size=page_size,
                session=session
            )
        
        # Reset page if out of bounds (happens when search changes)
        if result['total_pages'] > 0 and st.session_state.videos_dialog_page >= result['total_pages']:
            st.session_state.videos_dialog_page = 0
            with get_db_session() as session:
                result = VideoService.search_videos(
                    search_term=video_search,
                    show_archived=show_archived,
                    show_only_unassigned=show_only_unassigned,
                    page=0,
                    page_size=page_size,
                    session=session
                )
        
        # Show results info
        custom_info(f"Showing {len(result['videos'])} of {result['total_count']} videos (Page {result['page'] + 1} of {result['total_pages']})")
        
        # Pagination navigation
        if result['total_pages'] > 1:
            nav_col1, nav_col2, nav_col3 = st.columns([1, 1, 1])
            
            with nav_col1:
                if st.button("◀ Previous", disabled=(result['page'] == 0), use_container_width=True):
                    st.session_state.videos_dialog_page = max(0, result['page'] - 1)
                    st.rerun(scope="fragment")
            
            with nav_col2:
                page_options = [f"Page {i+1} / {result['total_pages']}" for i in range(result['total_pages'])]
                selected_page_display = st.selectbox(
                    "page_nav", 
                    page_options,
                    index=result['page'],
                    key="videos_dialog_page_select",
                    label_visibility="collapsed"
                )
                new_page = page_options.index(selected_page_display)
                if new_page != result['page']:
                    st.session_state.videos_dialog_page = new_page
                    st.rerun(scope="fragment")
            
            with nav_col3:
                if st.button("Next ▶", disabled=(result['page'] == result['total_pages'] - 1), use_container_width=True):
                    st.session_state.videos_dialog_page = min(result['total_pages'] - 1, result['page'] + 1)
                    st.rerun(scope="fragment")
        
        # Show the table
        if not result['videos'].empty:
            st.dataframe(result['videos'], use_container_width=True)
        else:
            st.warning("No videos match your search criteria.")
            
    except Exception as e:
        st.error(f"Error loading videos: {str(e)}")


@st.fragment
@st.dialog("📋 Projects Database Table", width="large")
def show_projects_table_dialog():
    """Show projects table in a dialog with search and pagination."""
    # Search and filter controls
    search_col1, search_col2 = st.columns([2, 1])
    with search_col1:
        project_search = st.text_input("🔍 Search projects", placeholder="Project name...", key="projects_dialog_project_search")
    with search_col2:
        show_archived = st.checkbox("Show archived", value=True, key="projects_dialog_show_archived")
    
    # Pagination controls
    page_size = st.selectbox("Projects per page", [25, 50, 100], index=1, key="projects_dialog_page_size")
    
    # Initialize page in session state
    if "projects_dialog_page" not in st.session_state:
        st.session_state.projects_dialog_page = 0
    
    try:
        with get_db_session() as session:
            # Search projects with pagination
            result = ProjectService.search_projects(
                search_term=project_search,
                show_archived=show_archived,
                page=st.session_state.projects_dialog_page,
                page_size=page_size,
                session=session
            )
        
        # Reset page if out of bounds (happens when search changes)
        if result['total_pages'] > 0 and st.session_state.projects_dialog_page >= result['total_pages']:
            st.session_state.projects_dialog_page = 0
            with get_db_session() as session:
                result = ProjectService.search_projects(
                    search_term=project_search,
                    show_archived=show_archived,
                    page=0,
                    page_size=page_size,
                    session=session
                )
        
        # Show results info
        custom_info(f"Showing {len(result['projects'])} of {result['total_count']} projects (Page {result['page'] + 1} of {result['total_pages']})")
        
        # Pagination navigation
        if result['total_pages'] > 1:
            nav_col1, nav_col2, nav_col3 = st.columns([1, 1, 1])
            
            with nav_col1:
                if st.button("◀ Previous", disabled=(result['page'] == 0), use_container_width=True):
                    st.session_state.projects_dialog_page = max(0, result['page'] - 1)
                    st.rerun(scope="fragment")
            
            with nav_col2:
                page_options = [f"Page {i+1} / {result['total_pages']}" for i in range(result['total_pages'])]
                selected_page_display = st.selectbox(
                    "page_nav", 
                    page_options,
                    index=result['page'],
                    key="projects_dialog_page_select",
                    label_visibility="collapsed"
                )
                new_page = page_options.index(selected_page_display)
                if new_page != result['page']:
                    st.session_state.projects_dialog_page = new_page
                    st.rerun(scope="fragment")
            
            with nav_col3:
                if st.button("Next ▶", disabled=(result['page'] == result['total_pages'] - 1), use_container_width=True):
                    st.session_state.projects_dialog_page = min(result['total_pages'] - 1, result['page'] + 1)
                    st.rerun(scope="fragment")
        
        # Show the table
        if not result['projects'].empty:
            st.dataframe(result['projects'], use_container_width=True)
        else:
            st.warning("No projects match your search criteria.")
            
    except Exception as e:
        st.error(f"Error loading projects: {str(e)}")


@st.fragment
@st.dialog("📋 Schemas Database Table", width="large")
def show_schemas_table_dialog():
    """Show schemas table in a dialog with search and pagination."""
    # Search and filter controls
    search_col1, search_col2 = st.columns([2, 1])
    with search_col1:
        schema_search = st.text_input("🔍 Search schemas", placeholder="Schema name...", key="schemas_dialog_schema_search")
    with search_col2:
        show_archived = st.checkbox("Show archived", key="schemas_dialog_show_archived")
    
    # Pagination controls
    page_size = st.selectbox("Schemas per page", [10, 20, 50], index=1, key="schemas_dialog_page_size")
    
    # Initialize page in session state
    if "schemas_dialog_page" not in st.session_state:
        st.session_state.schemas_dialog_page = 0
    
    try:
        with get_db_session() as session:
            # Search schemas with pagination
            result = SchemaService.search_schemas(
                search_term=schema_search,
                show_archived=show_archived,
                page=st.session_state.schemas_dialog_page,
                page_size=page_size,
                session=session
            )
        
        # Reset page if out of bounds (happens when search changes)
        if result['total_pages'] > 0 and st.session_state.schemas_dialog_page >= result['total_pages']:
            st.session_state.schemas_dialog_page = 0
            with get_db_session() as session:
                result = SchemaService.search_schemas(
                    search_term=schema_search,
                    show_archived=show_archived,
                    page=0,
                    page_size=page_size,
                    session=session
                )
        
        # Show results info
        custom_info(f"Showing {len(result['schemas'])} of {result['total_count']} schemas (Page {result['page'] + 1} of {result['total_pages']})")
        
        # Pagination navigation
        if result['total_pages'] > 1:
            nav_col1, nav_col2, nav_col3 = st.columns([1, 1, 1])
            
            with nav_col1:
                if st.button("◀ Previous", disabled=(result['page'] == 0), use_container_width=True):
                    st.session_state.schemas_dialog_page = max(0, result['page'] - 1)
                    st.rerun(scope="fragment")
            
            with nav_col2:
                page_options = [f"Page {i+1} / {result['total_pages']}" for i in range(result['total_pages'])]
                selected_page_display = st.selectbox(
                    "page_nav", 
                    page_options,
                    index=result['page'],
                    key="schemas_dialog_page_select",
                    label_visibility="collapsed"
                )
                new_page = page_options.index(selected_page_display)
                if new_page != result['page']:
                    st.session_state.schemas_dialog_page = new_page
                    st.rerun(scope="fragment")
            
            with nav_col3:
                if st.button("Next ▶", disabled=(result['page'] == result['total_pages'] - 1), use_container_width=True):
                    st.session_state.schemas_dialog_page = min(result['total_pages'] - 1, result['page'] + 1)
                    st.rerun(scope="fragment")
        
        # Show the table
        if not result['schemas'].empty:
            st.dataframe(result['schemas'], use_container_width=True)
        else:
            st.warning("No schemas match your search criteria.")
            
    except Exception as e:
        st.error(f"Error loading schemas: {str(e)}")


@st.fragment
@st.dialog("📋 Question Groups Database Table", width="large")
def show_question_groups_table_dialog():
    """Show question groups table in a dialog."""
    # Search and filter controls
    search_col1, search_col2 = st.columns([2, 1])
    with search_col1:
        search_term = st.text_input("🔍 Search groups", placeholder="Group name or description...", key="qgroups_dialog_group_search")
    with search_col2:
        show_archived = st.checkbox("Show archived", value=False, key="qgroups_dialog_show_archived")
    
    try:
        with get_db_session() as session:
            groups_df = QuestionGroupService.get_all_groups(session=session)
        
        # Apply filters
        filtered_groups = groups_df.copy()
        if not show_archived:
            filtered_groups = filtered_groups[~filtered_groups["Archived"]]
        
        if search_term:
            mask = (
                filtered_groups["Name"].str.contains(search_term, case=False, na=False) |
                filtered_groups["Description"].str.contains(search_term, case=False, na=False)
            )
            filtered_groups = filtered_groups[mask]
        
        custom_info(f"Showing {len(filtered_groups)} groups")
        
        if not filtered_groups.empty:
            st.dataframe(filtered_groups, use_container_width=True)
        else:
            st.warning("No groups match your search criteria.")
    except Exception as e:
        st.error(f"Error loading question groups: {str(e)}")


@st.fragment
@st.dialog("📋 Questions Database Table", width="large")
def show_questions_table_dialog():
    """Show questions table in a dialog with search and pagination."""

    # Search and filter controls
    search_col1, search_col2 = st.columns([2, 1])
    with search_col1:
        question_search = st.text_input("🔍 Search questions", placeholder="Question text...", key="questions_dialog_question_search")
    with search_col2:
        show_archived = st.checkbox("Show archived", value=False, key="questions_dialog_show_archived")
    
    # Pagination controls
    page_size = st.selectbox("Questions per page", [20, 50, 100], index=1, key="questions_dialog_page_size")
    
    # Initialize page in session state
    if "questions_dialog_page" not in st.session_state:
        st.session_state.questions_dialog_page = 0
    
    try:
        with get_db_session() as session:
            # Search questions with pagination
            result = QuestionService.search_questions(
                search_term=question_search,
                show_archived=show_archived,
                page=st.session_state.questions_dialog_page,
                page_size=page_size,
                session=session
            )
        
        # Reset page if out of bounds (happens when search changes)
        if result['total_pages'] > 0 and st.session_state.questions_dialog_page >= result['total_pages']:
            st.session_state.questions_dialog_page = 0
            with get_db_session() as session:
                result = QuestionService.search_questions(
                    search_term=question_search,
                    show_archived=show_archived,
                    page=0,
                    page_size=page_size,
                    session=session
                )
        
        # Show results info
        custom_info(f"Showing {len(result['questions'])} of {result['total_count']} questions (Page {result['page'] + 1} of {result['total_pages']})")
        
        # Pagination navigation
        if result['total_pages'] > 1:
            nav_col1, nav_col2, nav_col3 = st.columns([1, 1, 1])
            
            with nav_col1:
                if st.button("◀ Previous", disabled=(result['page'] == 0), use_container_width=True):
                    st.session_state.questions_dialog_page = max(0, result['page'] - 1)
                    st.rerun(scope="fragment")
            
            with nav_col2:
                page_options = [f"Page {i+1} / {result['total_pages']}" for i in range(result['total_pages'])]
                selected_page_display = st.selectbox(
                    "page_nav", 
                    page_options,
                    index=result['page'],
                    key="questions_dialog_page_select",
                    label_visibility="collapsed"
                )
                new_page = page_options.index(selected_page_display)
                if new_page != result['page']:
                    st.session_state.questions_dialog_page = new_page
                    st.rerun(scope="fragment")
            
            with nav_col3:
                if st.button("Next ▶", disabled=(result['page'] == result['total_pages'] - 1), use_container_width=True):
                    st.session_state.questions_dialog_page = min(result['total_pages'] - 1, result['page'] + 1)
                    st.rerun(scope="fragment")
        
        # Show the table
        if not result['questions'].empty:
            st.dataframe(result['questions'], use_container_width=True)
        else:
            st.warning("No questions match your search criteria.")
            
    except Exception as e:
        st.error(f"Error loading questions: {str(e)}")


@st.fragment
@st.dialog("📊 Project Groups Database Table", width="large")
def show_project_groups_table_dialog():
    """Show project groups table in a dialog."""
    # Search and filter controls
    search_col1, search_col2 = st.columns([2, 1])
    with search_col1:
        group_search = st.text_input("🔍 Search groups", placeholder="Group name or description...", key="pgroups_dialog_group_search")
    with search_col2:
        show_details = st.checkbox("Show project details", value=True, key="pgroups_dialog_show_details")
    
    # Pagination controls
    page_size = st.selectbox("Groups per page", [10, 20, 50], index=1, key="pgroups_dialog_page_size")
    
    # Initialize page in session state
    if "pgroups_dialog_page" not in st.session_state:
        st.session_state.pgroups_dialog_page = 0
    
    try:
        with get_db_session() as session:
            groups = ProjectGroupService.list_project_groups(session=session)
        
        if groups:
            # Build enhanced data
            group_data = []
            for group in groups:
                try:
                    with get_db_session() as session:
                        group_info = ProjectGroupService.get_project_group_by_id(group_id=int(group["id"]), session=session)
                    project_count = len(group_info["projects"])
                    project_names = [p["name"] for p in group_info["projects"][:3]]
                    project_preview = ", ".join(project_names)
                    if len(group_info["projects"]) > 3:
                        project_preview += f" (+{len(group_info['projects']) - 3} more)"
                    
                    group_data.append({
                        "ID": group["id"],
                        "Name": group["name"],
                        "Description": group["description"][:100] + "..." if len(str(group["description"])) > 100 else group["description"],
                        "Project Count": project_count,
                        "Sample Projects": project_preview,
                        "Created": group["created_at"]
                    })
                except Exception as e:
                    group_data.append({
                        "ID": group["id"],
                        "Name": group["name"],
                        "Description": group["description"],
                        "Project Count": "Error",
                        "Sample Projects": f"Error: {str(e)}",
                        "Created": group["created_at"]
                    })
            
            # Apply search filter
            if group_search:
                filtered_data = [
                    g for g in group_data 
                    if group_search.lower() in str(g["Name"]).lower() or 
                        group_search.lower() in str(g["Description"]).lower()
                ]
            else:
                filtered_data = group_data
            
            custom_info(f"Showing {len(filtered_data)} of {len(group_data)} total groups")
            
            if filtered_data:
                st.dataframe(pd.DataFrame(filtered_data), use_container_width=True)
            else:
                st.warning("No groups match your search criteria.")
        else:
            custom_info("No project groups exist yet.")
            
    except Exception as e:
        st.error(f"Error loading project groups: {str(e)}")


@st.fragment
@st.dialog("👥 Users Database Table", width="large")
def show_users_table_dialog():
    """Show users table in a dialog with search and filtering."""
    # Search and filter controls
    search_col1, search_col2, search_col3 = st.columns(3)
    with search_col1:
        user_search = st.text_input("🔍 Search users", placeholder="Name or email...", key="users_dialog_user_search")
    with search_col2:
        show_archived = st.checkbox("Show archived", value=True, key="users_dialog_show_archived")
    with search_col3:
        role_filter = st.selectbox("Filter by role", ["All", "admin", "human", "model"], key="users_dialog_role_filter")
    
    try:
        with get_db_session() as session:
            users_df = AuthService.get_all_users(session=session)
        
        # Apply filters
        filtered_users = users_df.copy()
        
        if not show_archived:
            filtered_users = filtered_users[~filtered_users["Archived"]]
        
        if role_filter != "All":
            filtered_users = filtered_users[filtered_users["Role"] == role_filter]
        
        if user_search:
            mask = (
                filtered_users["User ID"].str.contains(user_search, case=False, na=False) |
                filtered_users["Email"].str.contains(user_search, case=False, na=False)
            )
            filtered_users = filtered_users[mask]
        
        custom_info(f"Showing {len(filtered_users)} of {len(users_df)} total users")
        
        if not filtered_users.empty:
            st.dataframe(filtered_users, use_container_width=True)
        else:
            st.warning("No users match your search criteria.")
            
    except Exception as e:
        st.error(f"Error loading users: {str(e)}")

@st.fragment
@st.dialog("🔍 Users & Their Assignments", width="large")
def show_assignments_table_dialog():
    """Show users with their assignments with search and pagination in a dialog."""
    # Search and filter controls
    st.markdown("#### 🔍 Search & Filter")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        search_term = st.text_input("🔍 Search", placeholder="User, email, or project...", key="assign_dialog_search")
    
    with col2:
        status_filter = st.selectbox("Status", ["All", "Active", "Archived"], key="assign_dialog_status_filter")
    
    with col3:
        user_role_filter = st.selectbox("User Role", ["All", "admin", "human", "model"], key="assign_dialog_user_role_filter")
    
    with col4:
        project_role_filter = st.selectbox("Assignment Role", ["All", "annotator", "reviewer", "admin", "model"], key="assign_dialog_project_role_filter")
    
    # Pagination controls
    page_size = st.selectbox("Users per page", [5, 10, 20], index=1, key="assign_dialog_page_size")
    
    # Initialize page in session state
    if "assignments_dialog_page" not in st.session_state:
        st.session_state.assignments_dialog_page = 0
    
    try:
        with get_db_session() as session:
            # Use your updated search_assignments method
            result = AuthService.search_assignments(
                search_term=search_term,
                status_filter=status_filter,
                user_role_filter=user_role_filter,
                project_role_filter=project_role_filter,
                page=st.session_state.assignments_dialog_page,
                page_size=page_size,
                session=session
            )
        
        # Show results info
        if result['total_count'] > 0:
            custom_info(f"Found {result['total_count']} users with assignments (Page {result['page'] + 1} of {result['total_pages']})")
            
            # Pagination navigation
            if result['total_pages'] > 1:
                nav_col1, nav_col2, nav_col3 = st.columns([1, 1, 1])
                
                with nav_col1:
                    if st.button("◀ Previous", disabled=(result['page'] == 0), key="assign_dialog_prev", use_container_width=True):
                        st.session_state.assignments_dialog_page = max(0, result['page'] - 1)
                        st.rerun(scope="fragment")
                
                with nav_col2:
                    page_options = [f"Page {i+1} / {result['total_pages']}" for i in range(result['total_pages'])]
                    selected_page_display = st.selectbox(
                        "page_nav", 
                        page_options,
                        index=result['page'],
                        key="assign_dialog_page_select",
                        label_visibility="collapsed"
                    )
                    new_page = page_options.index(selected_page_display)
                    if new_page != result['page']:
                        st.session_state.assignments_dialog_page = new_page
                        st.rerun(scope="fragment")
                
                with nav_col3:
                    if st.button("Next ▶", disabled=(result['page'] == result['total_pages'] - 1), key="assign_dialog_next", use_container_width=True):
                        st.session_state.assignments_dialog_page = min(result['total_pages'] - 1, result['page'] + 1)
                        st.rerun(scope="fragment")
            
            # Use the ORIGINAL display function
            display_assignments_cards_dialog(result['user_assignments'])
            
        else:
            st.warning("No users with assignments match your search criteria.")
            
    except Exception as e:
        st.error(f"Error loading assignments: {str(e)}")


def display_assignments_cards_dialog(user_assignments: Dict):
    """Display assignments in optimized card format for dialog."""
    
    def get_user_role_emoji(role):
        """Get appropriate emoji for user role"""
        role_emojis = {"human": "👤", "admin": "👑", "model": "🤖"}
        return role_emojis.get(role, "❓")
    
    for user_id, user_data in user_assignments.items():
        status_color = COLORS['danger'] if user_data["is_archived"] else COLORS['primary']
        
        # Create project summary for preview
        total_projects = len(user_data["projects"])
        project_names = list(user_data["projects"].values())[:3]
        if total_projects > 3:
            sample_text = f"Projects: {', '.join([p['name'] for p in project_names])}... (+{total_projects-3} more)"
        else:
            sample_text = f"Projects: {', '.join([p['name'] for p in project_names])}"
        
        with st.container():
            st.markdown(f"""
            <div style="border: 2px solid {status_color}; border-radius: 10px; padding: 15px; margin: 10px 0; background: linear-gradient(135deg, #ffffff, #f8f9fa); box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div style="flex: 1;">
                        <h4 style="margin: 0; color: #9553FE;">👤 {user_data['name']}</h4>
                        <p style="margin: 5px 0; color: #6c757d;">📧 {user_data['email']}</p>
                        <p style="margin: 5px 0; color: #6c757d; font-size: 0.9rem; font-style: italic;">{sample_text}</p>
                    </div>
                    <div style="text-align: right;">
                        <span style="background: {status_color}; color: white; padding: 5px 10px; border-radius: 15px; font-weight: bold; font-size: 0.9rem;">
                            {'🗄️ Archived' if user_data['is_archived'] else '✅ Active'}
                        </span>
                        <br><br>
                        <span style="color: #495057; font-weight: bold;">
                            {get_user_role_emoji(user_data['user_role'].lower())} {user_data['user_role'].upper()}
                        </span>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Show detailed assignments in expander
            with st.expander(f"📁 View assignments ({total_projects} projects)", expanded=False):
                for project_id, project_info in user_data["projects"].items():
                    st.markdown(f"### 📁 {project_info['name']}")
                    
                    role_parts = []
                    for role, role_data in project_info["role_assignments"].items():
                        # Build role display
                        completion_emoji = "✅ " if (role != "admin" and role_data["completed_date"]) else ""
                        
                        # Build date info
                        if role == "admin":
                            date_part = f"({role_data['assigned_date'] or 'not set'})"
                        else:
                            assigned = role_data['assigned_date'] or 'not set'
                            completed = role_data['completed_date'] if role_data['completed_date'] else None
                            date_part = f"({assigned} → {completed})" if completed else f"({assigned})"
                        
                        # Weight info
                        weight_part = f" [weight: {role_data['user_weight']:.1f}]" if role_data['user_weight'] != 1.0 else ""
                        
                        archived_indicator = " 🗄️" if role_data["archived"] else ""
                        
                        role_parts.append(f"  • {completion_emoji}**{role.title()}** {date_part}{weight_part}{archived_indicator}")
                    
                    for role_part in role_parts:
                        st.markdown(role_part)
                    
                    st.markdown("---")
